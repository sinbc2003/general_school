"""시트 snapshot 추출 + XLSX 변환 회귀 테스트.

배경: google_integration/export.py의 _build_xlsx_from_sheet_sync가 yjs_state를
디코드하지 않고 빈 시트(제목만)를 만들면서 "성공"으로 표시하는 조용한 유실 버그가
있었다. export.py와 drive/backup.py가 공용 헬퍼(app/services/sheet_snapshot.py)를
쓰도록 통합한 뒤, 셀이 실제로 추출되는지 검증한다.

기존 test_google_export / test_drive_backup의 sheet 픽스처는 yjs_state가 비어
fallback 경로만 탔기에 이 버그를 잡지 못했다 — 본 테스트가 그 공백을 메운다.
"""

import io

from openpyxl import load_workbook

from app.services.sheet_snapshot import (
    build_xlsx_from_sheet,
    extract_sheet_snapshot,
    sheet_snapshot_to_xlsx,
)


def _make_yjs_state(snapshot):
    """프론트(SheetEditor.tsx)와 동일하게 Y.Map("sheet")["snapshot"]에 plain list 저장.

    SheetEditor는 yDoc.getMap("sheet") + yMap.set("snapshot", <plain JS array>).
    plain array는 Yjs에서 "Any"(JSON) 값 → pycrdt가 Python list로 복원.
    """
    from pycrdt import Doc, Map
    doc = Doc()
    m = doc.get("sheet", type=Map)
    m["snapshot"] = snapshot
    return doc.get_update()


SNAPSHOT = [
    {
        "name": "1반",
        "celldata": [
            {"r": 0, "c": 0, "v": {"v": "이름", "bl": 1}},
            {"r": 0, "c": 1, "v": {"v": "점수"}},
            {"r": 1, "c": 0, "v": "홍길동"},
            {"r": 1, "c": 1, "v": {"v": 95}},
        ],
    }
]


def test_extract_roundtrip_returns_list():
    snap = extract_sheet_snapshot(_make_yjs_state(SNAPSHOT))
    assert isinstance(snap, list), "pycrdt가 plain array를 list로 복원해야 함 (추출 전제)"
    assert snap[0]["name"] == "1반"


def test_xlsx_contains_cells_not_just_title():
    snap = extract_sheet_snapshot(_make_yjs_state(SNAPSHOT))
    wb = load_workbook(io.BytesIO(sheet_snapshot_to_xlsx(snap, "fallback")))
    ws = wb["1반"]
    assert ws["A1"].value == "이름"
    assert ws["B1"].value == "점수"
    assert ws["A2"].value == "홍길동"
    assert ws["B2"].value == 95
    assert ws["A1"].font.bold is True  # bl=1 서식 보존


def test_build_from_sheet_matches_backup_path():
    """export(build_xlsx_from_sheet)와 backup이 같은 셀을 만든다 (silent-loss 회귀 방지)."""
    state = _make_yjs_state(SNAPSHOT)

    class _FakeSheet:
        yjs_state = state
        title = "통합본"

    # backup.py는 공용 헬퍼를 별칭으로 재노출 — 같은 함수여야 함
    from app.modules.drive.backup import _extract_sheet_snapshot, _sheet_snapshot_to_xlsx
    assert _extract_sheet_snapshot is extract_sheet_snapshot
    assert _sheet_snapshot_to_xlsx is sheet_snapshot_to_xlsx

    export_ws = load_workbook(io.BytesIO(build_xlsx_from_sheet(_FakeSheet())))["1반"]
    assert export_ws["A2"].value == "홍길동"
    assert export_ws["B2"].value == 95


def test_empty_yjs_state_falls_back_to_title():
    wb = load_workbook(io.BytesIO(sheet_snapshot_to_xlsx(extract_sheet_snapshot(None), "빈시트")))
    ws = wb["빈시트"]
    assert ws["A1"].value == "빈시트"

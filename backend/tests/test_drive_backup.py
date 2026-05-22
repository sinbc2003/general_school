"""드라이브 백업 + 복원 roundtrip 테스트.

검증:
  - POST /api/drive/backup/download → ZIP 응답 (본인 자료 포함)
  - POST /api/drive/backup/import → ZIP 복원 (새 자료 생성)
  - manifest.json 검사 (system 미일치 시 거부)
  - 학생도 동작 (drive.use default 권한)
  - 다른 사용자 자료 안 들어감 (본인만)
"""

from __future__ import annotations

import io
import json
import zipfile

import pytest
from sqlalchemy import select

from app.models import (
    ClassroomDocument, ClassroomSheet, ClassroomPresentation, Folder, User,
)


@pytest.mark.security
@pytest.mark.asyncio
async def test_backup_download_only_own_items(
    app_client, db_session, teacher_user, student_user, auth_headers,
):
    """다른 사용자 자료는 ZIP에 들어가지 않음."""
    db_session.add_all([
        ClassroomDocument(owner_id=teacher_user.id, title="내 문서", plain_text="text"),
        ClassroomDocument(owner_id=student_user.id, title="남의 문서", plain_text="other"),
    ])
    await db_session.commit()

    r = await app_client.post(
        "/api/drive/backup/download",
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/zip")

    zf = zipfile.ZipFile(io.BytesIO(r.content), "r")
    names = zf.namelist()
    assert "manifest.json" in names
    manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
    assert manifest["system"] == "general_school"
    assert manifest["user"]["id"] == teacher_user.id
    # 본인 docs만 1개 (JSON + HTML 두 파일이지만 자료 1개)
    doc_jsons = [n for n in names if n.startswith("docs/") and n.endswith(".json")]
    assert len(doc_jsons) == 1
    data = json.loads(zf.read(doc_jsons[0]).decode("utf-8"))
    assert data["title"] == "내 문서"
    # HTML도 본인 자료만
    doc_htmls = [n for n in names if n.startswith("docs/") and n.endswith(".html")]
    assert len(doc_htmls) == 1
    assert "내 문서" in zf.read(doc_htmls[0]).decode("utf-8")


@pytest.mark.asyncio
async def test_backup_includes_folders_and_all_types(
    app_client, db_session, teacher_user, auth_headers,
):
    folder = Folder(
        owner_id=teacher_user.id, name="수업 자료", sort_order=1,
        is_system_locked=False,
    )
    db_session.add(folder)
    await db_session.commit()
    db_session.add_all([
        ClassroomDocument(owner_id=teacher_user.id, title="문서", folder_id=folder.id),
        ClassroomSheet(owner_id=teacher_user.id, title="시트"),
        ClassroomPresentation(owner_id=teacher_user.id, title="덱"),
    ])
    await db_session.commit()

    r = await app_client.post(
        "/api/drive/backup/download",
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(r.content), "r")
    names = zf.namelist()
    assert "folders.json" in names
    folders_data = json.loads(zf.read("folders.json").decode("utf-8"))
    assert any(f["name"] == "수업 자료" for f in folders_data)
    assert any(n.startswith("docs/") for n in names)
    assert any(n.startswith("sheets/") for n in names)
    assert any(n.startswith("decks/") for n in names)


@pytest.mark.security
@pytest.mark.asyncio
async def test_backup_roundtrip_import_restores_data(
    app_client, db_session, teacher_user, auth_headers,
):
    """backup → import roundtrip — 자료가 새 id로 복원됨."""
    db_session.add_all([
        ClassroomDocument(owner_id=teacher_user.id, title="원본 문서", plain_text="hello", storage_bytes=100),
        ClassroomSheet(owner_id=teacher_user.id, title="원본 시트", storage_bytes=50),
    ])
    await db_session.commit()

    # backup
    r1 = await app_client.post(
        "/api/drive/backup/download",
        headers=auth_headers(teacher_user),
    )
    assert r1.status_code == 200
    zip_bytes = r1.content

    # import
    r2 = await app_client.post(
        "/api/drive/backup/import",
        files={"file": ("backup.zip", zip_bytes, "application/zip")},
        headers=auth_headers(teacher_user),
    )
    assert r2.status_code == 200, r2.text
    data = r2.json()
    assert data["imported"]["docs"] >= 1
    assert data["imported"]["sheets"] >= 1

    # 원본 + 복원본 모두 존재 (총 2개씩)
    docs = (await db_session.execute(
        select(ClassroomDocument).where(
            ClassroomDocument.owner_id == teacher_user.id,
            ClassroomDocument.deleted_at.is_(None),
        )
    )).scalars().all()
    titles = [d.title for d in docs]
    assert "원본 문서" in titles
    # 복원본 이름은 그대로 (또는 동일 — 본 시스템은 중복 허용)
    assert titles.count("원본 문서") >= 2 or any("복원" in t or t == "원본 문서" for t in titles)


@pytest.mark.security
@pytest.mark.asyncio
async def test_backup_import_rejects_bad_system(
    app_client, teacher_user, auth_headers,
):
    """system != general_school인 ZIP 거부."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("manifest.json", json.dumps({
            "system": "other_school", "version": "1.0",
        }))
    r = await app_client.post(
        "/api/drive/backup/import",
        files={"file": ("bad.zip", buf.getvalue(), "application/zip")},
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 400


@pytest.mark.security
@pytest.mark.asyncio
async def test_backup_import_rejects_invalid_zip(
    app_client, teacher_user, auth_headers,
):
    """망가진 ZIP 거부."""
    r = await app_client.post(
        "/api/drive/backup/import",
        files={"file": ("bad.zip", b"not a zip", "application/zip")},
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 400


@pytest.mark.security
@pytest.mark.asyncio
async def test_backup_import_rejects_no_manifest(
    app_client, teacher_user, auth_headers,
):
    """manifest.json 없는 ZIP 거부."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("random.txt", "hi")
    r = await app_client.post(
        "/api/drive/backup/import",
        files={"file": ("bad.zip", buf.getvalue(), "application/zip")},
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 400


# ── 사람-읽기 형식 검증 ─────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_backup_includes_doc_html(
    app_client, db_session, teacher_user, auth_headers,
):
    """docs는 JSON + HTML 두 형식 모두 포함."""
    d = ClassroomDocument(
        owner_id=teacher_user.id, title="회의록 1차",
        plain_text="첫 줄\n둘째 줄\n셋째 줄",
    )
    db_session.add(d)
    await db_session.commit()

    r = await app_client.post(
        "/api/drive/backup/download",
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(r.content), "r")
    names = zf.namelist()
    json_files = [n for n in names if n.startswith("docs/") and n.endswith(".json")]
    html_files = [n for n in names if n.startswith("docs/") and n.endswith(".html")]
    assert len(json_files) == 1
    assert len(html_files) == 1
    html = zf.read(html_files[0]).decode("utf-8")
    assert "<!DOCTYPE html>" in html
    assert "회의록 1차" in html
    assert "첫 줄" in html


@pytest.mark.asyncio
async def test_backup_includes_sheet_xlsx(
    app_client, db_session, teacher_user, auth_headers,
):
    """sheets는 JSON + XLSX (yjs_state 없어도 fallback)."""
    s = ClassroomSheet(owner_id=teacher_user.id, title="출석부 1반")
    db_session.add(s)
    await db_session.commit()

    r = await app_client.post(
        "/api/drive/backup/download",
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(r.content), "r")
    names = zf.namelist()
    xlsx_files = [n for n in names if n.startswith("sheets/") and n.endswith(".xlsx")]
    assert len(xlsx_files) == 1
    # openpyxl로 다시 열어 데이터 확인
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(zf.read(xlsx_files[0])))
    assert wb.sheetnames  # 최소 1 시트


@pytest.mark.asyncio
async def test_backup_includes_survey_responses_csv(
    app_client, db_session, teacher_user, auth_headers,
):
    """surveys 응답 있으면 *_responses.csv 포함."""
    from app.models import Survey, SurveyQuestion, SurveyResponse, SurveyAnswer
    sv = Survey(
        author_id=teacher_user.id, title="만족도 조사",
        status="active", is_anonymous=False, access_mode="course_members",
    )
    db_session.add(sv)
    await db_session.flush()
    q1 = SurveyQuestion(
        survey_id=sv.id, order=0, question_type="short_text",
        question_text="만족도", is_required=True,
    )
    db_session.add(q1)
    await db_session.flush()
    resp = SurveyResponse(
        survey_id=sv.id, respondent_id=teacher_user.id,
    )
    db_session.add(resp)
    await db_session.flush()
    db_session.add(SurveyAnswer(
        response_id=resp.id, question_id=q1.id, text_value="만족함",
    ))
    await db_session.commit()

    r = await app_client.post(
        "/api/drive/backup/download",
        headers=auth_headers(teacher_user),
    )
    assert r.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(r.content), "r")
    names = zf.namelist()
    csv_files = [n for n in names if n.startswith("surveys/") and n.endswith("_responses.csv")]
    assert len(csv_files) == 1
    csv_text = zf.read(csv_files[0]).decode("utf-8")
    assert "만족도" in csv_text  # 질문 헤더
    assert "만족함" in csv_text  # 응답

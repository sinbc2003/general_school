"""사용자 일괄 등록 CSV (역할별)

자체 양식 — UTF-8 BOM. 헤더 첫 줄. **한글·영문 헤더 둘 다 인식**.

역할별 컬럼 (필수/선택):
  designated_admin: 이름*, 이메일*, 아이디*, 연락처
  teacher:           이름*, 이메일*, 아이디*, 연락처, 부서, 담당과목(개설과목 드롭다운)
  student:           이름*, 이메일*, 아이디*, 연락처, 학년, 반, 번호

초기 비밀번호 = 연락처(숫자만, '-' 없이). 연락처 없으면 settings.DEFAULT_USER_PASSWORD. must_change_password=True.
(비밀번호 컬럼은 폐지 — 연락처가 곧 초기 비번. 단 호환을 위해 password 컬럼이 있으면 그 값을 우선 사용)
이메일/아이디 중복은 행 단위 실패로 보고.
"""

import csv
import io

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth import hash_password
from app.core.config import settings
from app.core.quota import assign_default_quota
from app.models.user import User


# 한글 헤더 → 표준 영문 키 매핑 (한국 학교 친화)
COLUMN_ALIASES: dict[str, str] = {
    # 공통
    "이름": "name", "성명": "name", "name": "name",
    "이메일": "email", "메일": "email", "email": "email", "e-mail": "email",
    "아이디": "username", "ID": "username", "id": "username", "username": "username", "유저네임": "username",
    "비밀번호": "password", "패스워드": "password", "password": "password", "비번": "password",
    "연락처": "phone", "전화번호": "phone", "휴대폰": "phone", "핸드폰": "phone", "전화": "phone", "phone": "phone",
    # 교사
    "부서": "department", "department": "department", "소속": "department",
    "담당과목": "teaching_subject", "과목": "teaching_subject", "subject": "teaching_subject", "teaching_subject": "teaching_subject",
    # 학생
    "학년": "grade", "grade": "grade",
    "반": "class_number", "학급": "class_number", "class_number": "class_number", "class": "class_number",
    "번호": "student_number", "출석번호": "student_number", "student_number": "student_number", "no": "student_number",
}


CSV_TEMPLATES: dict[str, list[str]] = {
    "designated_admin": ["이름", "이메일", "아이디", "연락처"],
    "teacher":          ["이름", "이메일", "아이디", "연락처", "부서", "담당과목"],
    "student":          ["이름", "이메일", "아이디", "연락처",
                         "학년", "반", "번호"],
}

import re as _re
_PHONE_NONDIGIT = _re.compile(r"\D")


def _phone_digits(phone: str | None) -> str | None:
    """전화번호 → 숫자만 (초기 비밀번호용). 숫자 없으면 None. (semester_import와 동일 규칙)"""
    if not phone:
        return None
    d = _PHONE_NONDIGIT.sub("", phone)
    return d or None

# 표준 영문 키 기준 (한글 → 영문 변환 후 검사)
REQUIRED: dict[str, set[str]] = {
    "designated_admin": {"name", "email", "username"},
    "teacher":          {"name", "email", "username"},
    "student":          {"name", "email", "username"},
}


def _normalize_row(row: dict) -> dict:
    """한글/영문 헤더를 표준 영문 키로 변환."""
    out = {}
    for k, v in row.items():
        if k is None:
            continue
        norm = COLUMN_ALIASES.get(k.strip(), k.strip())
        out[norm] = v
    return out


def template_csv(role: str) -> str:
    """한글 헤더 CSV + 예시 + 설명 셀 (B2/C2 같은 비활용 셀에 안내).

    UTF-8 BOM 포함 — Excel·LibreOffice에서 한글 안 깨짐.
    """
    if role not in CSV_TEMPLATES:
        raise ValueError(f"unknown role: {role}")
    cols = CSV_TEMPLATES[role]
    BOM = "﻿"
    header = ",".join(cols)

    if role == "student":
        example = "홍길동,gildong@school.local,10101,01012345678,1,1,1"
        descs = [
            "# 이름 필수 (실명)",
            "# 이메일 필수 (학교/개인 메일)",
            "# 아이디 필수 (학번 권장: 학년+반+번호 5자리 예 10101)",
            "# 연락처 = 초기 비밀번호 ('-' 없이 숫자만 입력). 첫 로그인 시 강제 변경",
            "# 학년 1·2·3",
            "# 반 숫자",
            "# 번호 출석번호",
        ]
    elif role == "teacher":
        example = "김선생,kim@school.local,kimss,01012345678,수학과,수학"
        descs = [
            "# 이름 필수 (실명)",
            "# 이메일 필수 (학교 메일)",
            "# 아이디 필수 (영문+숫자 8자 이상 권장)",
            "# 연락처 = 초기 비밀번호 ('-' 없이 숫자만 입력). 첫 로그인 시 강제 변경",
            "# 부서명 (마법사 2단계에서 등록한 부서명과 정확히 일치)",
            "# 담당 과목 — 개설 과목 목록에서 선택(드롭다운). 복수 과목은 등록 후 '내 정보'에서 추가",
        ]
    else:
        example = "지정관리자,da@school.local,da01,01012345678"
        descs = [
            "# 이름 필수",
            "# 이메일 필수",
            "# 아이디 필수",
            "# 연락처 = 초기 비밀번호 ('-' 없이 숫자만 입력)",
        ]

    # 헤더 + 예시 행 + 빈 셀에 설명 (Excel에서 #로 시작하는 셀은 데이터 무시)
    # 각 컬럼 설명을 3행에 배치 (영향 없는 행)
    desc_row = ",".join(descs) if len(descs) == len(cols) else ""
    return BOM + header + "\n" + example + "\n" + desc_row + "\n"


# 부서 드롭다운에 항상 포함되는 기본 항목 (직책/부서 무관 케이스)
FIXED_DEPARTMENTS: list[str] = ["교장", "교감", "행정실", "기타"]


def template_xlsx(role: str, dept_names: list[str] | None = None,
                  subject_names: list[str] | None = None) -> bytes:
    """역할별 xlsx 템플릿. teacher면 '부서'·'담당과목' 열에 드롭다운(DataValidation).

    dept_names: 부서 드롭다운 목록 (DB 등록 부서 + 고정 항목). None이면 드롭다운 없음.
    subject_names: 담당과목 드롭다운 목록 (현재 학기 개설 과목). None이면 드롭다운 없음.
    목록은 숨김 시트에 넣어 255자 제한을 회피.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill

    if role not in CSV_TEMPLATES:
        raise ValueError(f"unknown role: {role}")
    cols = CSV_TEMPLATES[role]

    wb = Workbook()
    ws = wb.active
    ws.title = "등록"
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A4A6E")

    examples = {
        "teacher": ["김선생", "kim@school.local", "kimss", "01012345678", "수학과", "수학"],
        "student": ["홍길동", "gildong@school.local", "10101", "01012345678", "1", "1", "1"],
        "designated_admin": ["지정관리자", "da@school.local", "da01", "01012345678"],
    }
    ws.append(examples.get(role, []))
    for c in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    if role == "teacher" and "부서" in cols and dept_names:
        ws_list = wb.create_sheet("_부서목록")
        for i, d in enumerate(dept_names, 1):
            ws_list.cell(row=i, column=1, value=d)
        ws_list.sheet_state = "hidden"
        col_idx = cols.index("부서") + 1
        letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=f"_부서목록!$A$1:$A${len(dept_names)}",
            allow_blank=True,
        )
        dv.prompt = "목록에서 부서/직책을 선택하세요"
        dv.promptTitle = "부서 선택"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}1000")

    if role == "teacher" and "담당과목" in cols and subject_names:
        ws_subj = wb.create_sheet("_과목목록")
        for i, s in enumerate(subject_names, 1):
            ws_subj.cell(row=i, column=1, value=s)
        ws_subj.sheet_state = "hidden"
        col_idx = cols.index("담당과목") + 1
        letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=f"_과목목록!$A$1:$A${len(subject_names)}",
            allow_blank=True,
        )
        dv.prompt = "개설 과목 목록에서 선택하세요"
        dv.promptTitle = "담당 과목 선택"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}1000")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_rows(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    """업로드 파일 파싱 → (헤더, row dict 목록). xlsx(PK 매직) 또는 CSV 자동 인식."""
    if file_bytes[:2] == b"PK":  # xlsx (zip) 매직바이트
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return [], []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        out: list[dict] = []
        for r in rows[1:]:
            d: dict = {}
            for idx, h in enumerate(header):
                if not h:
                    continue
                val = r[idx] if idx < len(r) else None
                d[h] = str(val).strip() if val is not None else ""
            out.append(d)
        return header, out
    # CSV (UTF-8 BOM 허용)
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    header = [f.strip() for f in (reader.fieldnames or [])]
    return header, [dict(r) for r in reader]


def _to_int(v: str) -> int | None:
    v = (v or "").strip()
    return int(v) if v else None


async def import_users_csv(
    db: AsyncSession, role: str, file_bytes: bytes,
    granted_by_user_id: int, dry_run: bool = False,
) -> dict:
    """CSV 일괄 등록. role 외의 컬럼은 무시.
    반환: {ok_count, errors: [{row, error}], dry_run}
    """
    if role not in CSV_TEMPLATES:
        raise ValueError(f"unknown role: {role}")

    header, rows = _read_rows(file_bytes)
    # 헤더를 표준 영문 키로 변환 (한글 헤더 지원)
    raw_fields = [f.strip() for f in header]
    normalized_fields = {COLUMN_ALIASES.get(f, f) for f in raw_fields}
    missing = REQUIRED[role] - normalized_fields
    if missing:
        # 한글 표현으로 에러 메시지
        ko_map = {"name": "이름", "email": "이메일", "username": "아이디"}
        ko_missing = [ko_map.get(m, m) for m in sorted(missing)]
        return {
            "ok_count": 0,
            "errors": [{"row": 1, "error": f"필수 컬럼 누락: {ko_missing}"}],
            "dry_run": dry_run,
        }

    # 현재 학기 + 개설 과목 (담당과목 검증·enrollment 생성용)
    import json as _json
    from app.models.timetable import Semester, SemesterEnrollment
    cur_sem = (await db.execute(select(Semester).where(Semester.is_current == True))).scalar_one_or_none()
    valid_subjects: set[str] = set()
    if cur_sem and cur_sem.subjects:
        try:
            valid_subjects = {str(x).strip() for x in _json.loads(cur_sem.subjects) if str(x).strip()}
        except Exception:
            valid_subjects = set()

    ok_count = 0
    errors: list[dict] = []
    new_users: list[User] = []
    teaching_subject_by_user: dict[int, str] = {}  # id(User) → 담당과목 (교사)
    seen_emails: set[str] = set()
    seen_usernames: set[str] = set()

    # 기존 DB 중복 한 번에 조회
    existing_emails = set(
        (await db.execute(select(User.email))).scalars().all()
    )
    existing_usernames = set(
        u for u in (await db.execute(select(User.username))).scalars().all() if u
    )

    for i, row in enumerate(rows, start=2):
        # 한글 헤더 row를 영문 키로 변환
        row = _normalize_row(row)

        # 주석 행(# 으로 시작) skip — 템플릿의 설명 행 무시
        first_val = (row.get("name") or "").strip()
        if first_val.startswith("#"):
            continue
        try:
            name = first_val
            email = (row.get("email") or "").strip().lower()
            username = (row.get("username") or "").strip()
            phone = (row.get("phone") or "").strip() or None
            # 초기 비번: 명시 비번 → 연락처(숫자만) → 기본값
            password = (
                (row.get("password") or "").strip()
                or _phone_digits(phone)
                or settings.DEFAULT_USER_PASSWORD
            )

            if not name or not email or not username:
                raise ValueError("이름/이메일/아이디 필수")
            if email in existing_emails or email in seen_emails:
                raise ValueError(f"이미 등록된 이메일: {email}")
            if username in existing_usernames or username in seen_usernames:
                raise ValueError(f"이미 등록된 아이디: {username}")
            seen_emails.add(email)
            seen_usernames.add(username)

            kwargs = dict(
                name=name, email=email, username=username,
                password_hash=hash_password(password),
                role=role, status="approved",
                phone=phone,
                must_change_password=True,
            )
            tsub = None
            if role == "teacher":
                kwargs["department"] = (row.get("department") or "").strip() or None
                tsub = (row.get("teaching_subject") or "").strip() or None
                if tsub and valid_subjects and tsub not in valid_subjects:
                    raise ValueError(f"등록되지 않은 과목: {tsub} (개설 과목 목록에서 선택)")
            elif role == "student":
                kwargs["grade"] = _to_int(row.get("grade") or "")
                kwargs["class_number"] = _to_int(row.get("class_number") or "")
                kwargs["student_number"] = _to_int(row.get("student_number") or "")

            u = User(**kwargs)
            new_users.append(u)
            if tsub:
                teaching_subject_by_user[id(u)] = tsub
            ok_count += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    if not dry_run and new_users:
        for u in new_users:
            assign_default_quota(u)
            db.add(u)
        await db.flush()
        # 현재 학기 명단(enrollment) 자동 등록 — create_user와 동일 정책(CSV도 명단에 들어가야
        # 클래스룸/시간표 매칭됨). 교사는 담당과목(teaching_subjects)도 함께 저장.
        if cur_sem:
            try:
                for u in new_users:
                    dup = (await db.execute(select(SemesterEnrollment).where(
                        SemesterEnrollment.semester_id == cur_sem.id,
                        SemesterEnrollment.user_id == u.id,
                    ))).scalar_one_or_none()
                    if not dup:
                        db.add(SemesterEnrollment(
                            semester_id=cur_sem.id, user_id=u.id, role=u.role, status="active",
                            grade=u.grade, class_number=u.class_number, student_number=u.student_number,
                            department=u.department, phone=u.phone,
                            teaching_subjects=teaching_subject_by_user.get(id(u)),
                        ))
                await db.flush()
            except Exception:
                pass

    return {"ok_count": ok_count, "errors": errors, "dry_run": dry_run}

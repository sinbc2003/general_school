"""협업 시트(fortune-sheet) snapshot 추출 + XLSX 변환 — 공용 SSOT.

ClassroomSheet.yjs_state(바이너리 CRDT)에는 fortune-sheet workbook JSON이
Y.Map("sheet")의 "snapshot" key에 저장된다 (SheetEditor가 HocuspocusProvider
name="sheet-{id}"로 동기화). 이 모듈은 그 snapshot을 디코드하고 openpyxl XLSX로
변환하는 로직을 한 곳에 모은다.

- drive/backup.py        — ZIP 백업의 사람-읽기 .xlsx 생성
- google_integration/export.py — Google Sheets export 변환

※ 이전에는 export.py가 자체 stub(`pass`)으로 셀을 디코드하지 않아 빈 시트만
   만들면서 "성공"으로 표시하는 조용한 유실 버그가 있었다. 두 경로가 같은 헬퍼를
   쓰도록 통합해 재발을 막는다.

호출부가 동기(openpyxl·pycrdt는 sync)이므로 caller가 asyncio.to_thread로 감싼다.
"""

from __future__ import annotations

import io


def extract_sheet_snapshot(yjs_state: bytes | None) -> list[dict] | None:
    """sheets의 yjs_state에서 fortune-sheet snapshot(list[dict]) 추출.

    Y.Map("sheet") 안 "snapshot" key에 fortune-sheet workbook JSON 저장됨.
    pycrdt로 디코드. 실패/없음 시 None.
    """
    if not yjs_state:
        return None
    try:
        from pycrdt import Doc, Map
        doc = Doc()
        doc.apply_update(yjs_state)
        # Y.Map("sheet") — SheetEditor에서 사용 (HocuspocusProvider name="sheet-{id}")
        sheet_map = doc.get("sheet", type=Map)
        snap = sheet_map.get("snapshot")
        if isinstance(snap, list):
            return snap
        return None
    except Exception:
        return None


def sheet_snapshot_to_xlsx(snapshot: list[dict] | None, fallback_title: str) -> bytes:
    """fortune-sheet snapshot → openpyxl XLSX bytes.

    snapshot이 None/빈 값이면 제목만 담은 fallback 시트를 만든다(셀 데이터 없음 안내 포함).
    """
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    if not snapshot:
        ws = wb.create_sheet(title=(fallback_title or "Sheet1")[:31])
        ws["A1"] = fallback_title or "데이터 없음"
        ws["A2"] = "원본은 본 시스템 yjs_state에 있고 외부 변환이 불완전합니다."
    else:
        for sheet_data in snapshot:
            ws_name = (sheet_data.get("name") or "Sheet")[:31]
            # 동일 시트명 충돌 회피
            counter = 0
            unique = ws_name
            while unique in wb.sheetnames:
                counter += 1
                unique = f"{ws_name[:28]}_{counter}"
            ws = wb.create_sheet(title=unique)
            cells = sheet_data.get("celldata", [])
            for c in cells:
                # pycrdt가 Yjs "Any"(JSON) 숫자를 float로 복원하므로 인덱스는 int 강제.
                # (openpyxl의 ws.cell은 float 행/열 인덱스에서 TypeError)
                r = int(c.get("r") or 0)
                col = int(c.get("c") or 0)
                v = c.get("v", {})
                val = v.get("v") if isinstance(v, dict) else v
                if val is None:
                    continue
                # 정수형 float(95.0)는 정수로 정규화 — 셀에 95.0 대신 95 표시
                if isinstance(val, float) and val.is_integer():
                    val = int(val)
                cell = ws.cell(row=r + 1, column=col + 1, value=val)
                # 기본 서식 (굵게, 폰트 색상은 fortune-sheet의 ct/bl/fc 참조)
                if isinstance(v, dict):
                    try:
                        from openpyxl.styles import Font
                        bold = bool(v.get("bl"))
                        color = v.get("fc")
                        if bold or color:
                            cell.font = Font(
                                bold=bold,
                                color=(color.lstrip("#") if isinstance(color, str) else None),
                            )
                    except Exception:
                        pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_xlsx_from_sheet(sheet) -> bytes:
    """ClassroomSheet → XLSX bytes (snapshot 추출 + 변환 결합). sync.

    caller가 asyncio.to_thread(build_xlsx_from_sheet, sheet)로 감싼다.
    """
    snapshot = extract_sheet_snapshot(getattr(sheet, "yjs_state", None))
    return sheet_snapshot_to_xlsx(snapshot, getattr(sheet, "title", None) or "제목 없음")

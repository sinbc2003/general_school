"""엑셀 서비스 — 사용자 템플릿 생성, 파싱, 내보내기"""

import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.user import User

VALID_ROLES = {"teacher", "staff", "student"}
HEADER_FILL = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = [
    ("이름", 15),
    ("이메일", 30),
    ("역할 (teacher/staff/student)", 28),
    ("초기 비밀번호 (선택)", 22),
    ("학년 (학생 필수)", 16),
    ("반 (선택)", 10),
    ("번호 (선택)", 10),
    ("부서 (교사/직원)", 18),
]


def generate_user_template() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "사용자 등록"

    # 헤더
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    # 예시 데이터
    examples = [
        ["홍길동", "hong@school.kr", "student", "", 2, 3, 15, ""],
        ["김교사", "kim@school.kr", "teacher", "", "", "", "", "수학과"],
        ["박직원", "park@school.kr", "staff", "", "", "", "", "행정실"],
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(color="999999")

    # 안내 시트
    info = wb.create_sheet("안내사항")
    info.append(["사용자 일괄 등록 안내"])
    info.append([])
    info.append(["1. '사용자 등록' 시트에 데이터를 입력하세요."])
    info.append(["2. 이름과 이메일은 필수입니다."])
    info.append(["3. 역할: teacher(교사), staff(직원), student(학생) 중 선택"])
    info.append(["4. 초기 비밀번호를 비워두면 기본 비밀번호가 설정됩니다."])
    info.append(["5. 학생의 경우 학년은 필수입니다."])
    info.append(["6. 등록 후 첫 로그인 시 비밀번호 변경이 필요합니다."])
    info.append([])
    info.append(["※ 예시 데이터는 삭제하고 실제 데이터를 입력하세요."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def parse_user_excel(
    content: bytes, db: AsyncSession
) -> tuple[list[dict], list[dict]]:
    """엑셀 파일을 파싱하여 유효한 행과 오류 목록 반환"""
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    if not ws:
        return [], [{"row": 0, "field": "", "message": "시트를 찾을 수 없습니다"}]

    # 기존 이메일 조회
    result = await db.execute(select(User.email))
    existing_emails = set(result.scalars().all())

    valid_rows: list[dict] = []
    errors: list[dict] = []
    seen_emails: set[str] = set()

    email_re = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        name = str(row[0]).strip() if row[0] else ""
        email = str(row[1]).strip() if row[1] else ""
        role = str(row[2]).strip().lower() if row[2] else ""
        password = str(row[3]).strip() if row[3] else ""
        grade = row[4] if row[4] else None
        class_num = row[5] if row[5] else None
        student_num = row[6] if row[6] else None
        department = str(row[7]).strip() if row[7] else None

        # 유효성 검사
        row_errors = []

        if not name or len(name) < 2:
            row_errors.append({"row": row_idx, "field": "이름", "message": "이름은 2자 이상"})

        if not email or not email_re.match(email):
            row_errors.append({"row": row_idx, "field": "이메일", "message": "유효하지 않은 이메일"})
        elif email in existing_emails:
            row_errors.append({"row": row_idx, "field": "이메일", "message": f"이미 등록된 이메일: {email}"})
        elif email in seen_emails:
            row_errors.append({"row": row_idx, "field": "이메일", "message": f"파일 내 중복: {email}"})

        if role not in VALID_ROLES:
            row_errors.append({"row": row_idx, "field": "역할", "message": f"유효하지 않은 역할: {role}"})

        if role == "student":
            if not grade:
                row_errors.append({"row": row_idx, "field": "학년", "message": "학생은 학년 필수"})
            else:
                try:
                    grade = int(grade)
                    if grade not in (1, 2, 3):
                        row_errors.append({"row": row_idx, "field": "학년", "message": "학년은 1-3"})
                except (ValueError, TypeError):
                    row_errors.append({"row": row_idx, "field": "학년", "message": "숫자를 입력하세요"})

        if password and len(password) < 8:
            row_errors.append({"row": row_idx, "field": "비밀번호", "message": "비밀번호 8자 이상"})

        if row_errors:
            errors.extend(row_errors)
        else:
            seen_emails.add(email)
            data = {
                "name": name,
                "email": email,
                "role": role,
                "password": password if password else None,
                "grade": int(grade) if grade else None,
                "class_number": int(class_num) if class_num else None,
                "student_number": int(student_num) if student_num else None,
                "department": department if department else None,
            }
            valid_rows.append(data)

    return valid_rows, errors


def generate_user_export(users: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "사용자 목록"

    headers = ["ID", "이름", "이메일", "역할", "상태", "학년", "반", "번호", "부서", "등록일"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, u in enumerate(users, 2):
        vals = [
            u.id, u.name, u.email, u.role, u.status,
            u.grade, u.class_number, u.student_number,
            u.department,
            u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

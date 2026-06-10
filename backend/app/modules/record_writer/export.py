"""생기부 엑셀 내보내기 — NEIS 붙여넣기용.

GET /projects/{pid}/export.xlsx
열: 학번 | 이름 | (항목들 generated_text) | 최종 종합(final_text)
openpyxl은 동기 CPU-bound → asyncio.to_thread (event loop 비차단).
"""

import asyncio
import io
import urllib.parse

from fastapi import Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_action
from app.core.database import get_db
from app.core.permissions import require_permission
from app.models.student_record_project import (
    RecordCell,
    RecordColumn,
    RecordProjectStudent,
)
from app.models.user import User
from app.modules.record_writer._helpers import get_owned_project
from app.modules.record_writer.router import router


def _student_no(u: User) -> str:
    if u.grade and u.class_number and u.student_number:
        return f"{u.grade}{u.class_number:02d}{u.student_number:02d}"
    return ""


def _build_xlsx_sync(headers: list[str], rows: list[list[str]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "생활기록부"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    # 본문 열 폭 + 줄바꿈
    for idx in range(3, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 60
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    wrap = Alignment(wrap_text=True, vertical="top")
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = wrap
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/projects/{pid}/export.xlsx")
async def export_xlsx(
    pid: int,
    user: User = Depends(require_permission("record.project.view")),
    db: AsyncSession = Depends(get_db),
):
    p = await get_owned_project(db, user, pid)

    srows = (
        await db.execute(
            select(RecordProjectStudent, User)
            .join(User, User.id == RecordProjectStudent.student_id)
            .where(RecordProjectStudent.project_id == pid)
            .order_by(RecordProjectStudent.display_order)
        )
    ).all()
    cols = (
        await db.execute(
            select(RecordColumn).where(RecordColumn.project_id == pid)
            .order_by(RecordColumn.display_order)
        )
    ).scalars().all()
    cells = {
        (c.column_id, c.student_id): c
        for c in (
            await db.execute(select(RecordCell).where(RecordCell.project_id == pid))
        ).scalars().all()
    }

    headers = ["학번", "이름"] + [c.name for c in cols] + ["최종 종합"]
    rows: list[list[str]] = []
    for rps, u in srows:
        row = [_student_no(u), u.name]
        for c in cols:
            cell = cells.get((c.id, rps.student_id))
            row.append((cell.generated_text if cell else "") or "")
        row.append(rps.final_text or "")
        rows.append(row)

    data = await asyncio.to_thread(_build_xlsx_sync, headers, rows)
    await log_action(
        db, user, "record.export",
        detail=f"생기부 #{pid} 엑셀 내보내기 ({len(rows)}명)",
        is_sensitive=True,
    )
    fname = urllib.parse.quote(f"생기부_{p.name}.xlsx")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )

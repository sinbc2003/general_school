"""내 드라이브 백업 — 학교 이동 시 사용자 자료 전체 export.

옵션:
  1. ZIP 다운로드: 자료 5종 + 폴더 트리 + manifest.json
     - docs/sheets/decks/surveys: JSON(메타 + yjs_state base64 + plain_text)
     - hwps: 원본 file
     - manifest.json: 자료 list + 폴더 트리 + 사용자 정보
  2. Google Drive 일괄: docs/sheets만 우선 (google_integration/export.py 활용)

ZIP 구조:
  my-drive-backup-{user}-{yyyymmdd}.zip
    manifest.json
    folders/
      (folder tree in JSON)
    docs/
      {id}.json
    sheets/
      {id}.json
    decks/
      {id}.json
      slides_{id}.json  (ClassroomSlide rows 별도)
    surveys/
      {id}.json
      responses_{id}.json
    hwps/
      {id}/{원본 파일명}
      {id}.json  (메타)

자료 본문은 사람-읽기 변환(HTML/XLSX) 미포함 — 본 시스템 재import용 + plain_text fallback.
"""

from __future__ import annotations

import asyncio
import base64
import io
import json
import secrets
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_action
from app.core.database import get_db
from app.core.files import ensure_dir_async, write_bytes_async
from app.core.permissions import require_permission
from app.core.quota import check_quota, consume_quota
from app.core.upload import POLICY_BACKUP, validate_upload
from app.models import (
    ClassroomDocument, ClassroomHwp, ClassroomPresentation, ClassroomSheet,
    ClassroomSlide, Folder, Survey, SurveyAnswer, SurveyQuestion, SurveyResponse, User,
)
from app.core.files import DEFAULT_STORAGE_ROOT
from app.modules.drive.router import router

# settings.STORAGE_ROOT 기반 (Phase 2-Q 통합).
STORAGE_ROOT = DEFAULT_STORAGE_ROOT


def _serialize_folder(f: Folder) -> dict[str, Any]:
    return {
        "id": f.id,
        "parent_id": f.parent_id,
        "name": f.name,
        "auto_kind": f.auto_kind,
        "semester_id": f.semester_id,
        "source_kind": f.source_kind,
        "source_id": f.source_id,
        "sort_order": f.sort_order,
        "is_system_locked": f.is_system_locked,
        "created_at": f.created_at.isoformat() if f.created_at else None,
    }


def _serialize_doc(d: ClassroomDocument) -> dict[str, Any]:
    return {
        "id": d.id,
        "title": d.title,
        "course_id": d.course_id,
        "folder_id": d.folder_id,
        "access_mode": d.access_mode,
        "plain_text": d.plain_text,
        "yjs_state_base64": base64.b64encode(d.yjs_state).decode("ascii") if d.yjs_state else None,
        "storage_bytes": d.storage_bytes,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _serialize_sheet(s: ClassroomSheet) -> dict[str, Any]:
    return {
        "id": s.id,
        "title": s.title,
        "course_id": s.course_id,
        "folder_id": s.folder_id,
        "access_mode": s.access_mode,
        "settings": s.settings,
        "yjs_state_base64": base64.b64encode(s.yjs_state).decode("ascii") if s.yjs_state else None,
        "storage_bytes": s.storage_bytes,
        "source_survey_id": s.source_survey_id,
        "created_at": s.created_at.isoformat() if s.created_at else None,
        "updated_at": s.updated_at.isoformat() if s.updated_at else None,
    }


def _serialize_deck(p: ClassroomPresentation) -> dict[str, Any]:
    return {
        "id": p.id,
        "title": p.title,
        "course_id": p.course_id,
        "folder_id": p.folder_id,
        "access_mode": p.access_mode,
        "settings": p.settings,
        "yjs_state_base64": base64.b64encode(p.yjs_state).decode("ascii") if p.yjs_state else None,
        "storage_bytes": p.storage_bytes,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }


def _serialize_slide(sl: ClassroomSlide) -> dict[str, Any]:
    return {
        "id": sl.id,
        "presentation_id": sl.presentation_id,
        "order": sl.order,
        "title": sl.title,
        "plain_text": sl.plain_text,
        "settings": sl.settings,
    }


def _serialize_survey(sv: Survey, questions: list[SurveyQuestion]) -> dict[str, Any]:
    return {
        "id": sv.id,
        "title": sv.title,
        "description": sv.description,
        "course_id": sv.course_id,
        "folder_id": sv.folder_id,
        "status": sv.status,
        "is_anonymous": sv.is_anonymous,
        "access_mode": sv.access_mode,
        "questions": [
            {
                "id": q.id, "order": q.order, "question_type": q.question_type,
                "question_text": q.question_text, "is_required": q.is_required,
                "options": q.options,
            }
            for q in questions
        ],
        "created_at": sv.created_at.isoformat() if sv.created_at else None,
    }


def _serialize_hwp(h: ClassroomHwp) -> dict[str, Any]:
    return {
        "id": h.id,
        "title": h.title,
        "course_id": h.course_id,
        "folder_id": h.folder_id,
        "access_mode": h.access_mode,
        "file_path": h.file_path,
        "file_format": h.file_format,
        "storage_bytes": h.storage_bytes,
        "created_at": h.created_at.isoformat() if h.created_at else None,
    }


def _safe_name(s: str | None) -> str:
    """ZIP 안 파일명용. 윈도우 호환 (특수 문자 _)."""
    base = (s or "untitled").strip()
    for ch in '<>:"/\\|?*\n\r\t':
        base = base.replace(ch, "_")
    return base[:80] or "untitled"


# ─────────────────────────────────────────────────────────────────────────────
# 사람-읽기 변환 (XLSX / CSV / HTML)
# ─────────────────────────────────────────────────────────────────────────────


def _extract_sheet_snapshot(yjs_state: bytes | None) -> list[dict] | None:
    """sheets의 yjs_state에서 fortune-sheet snapshot 추출.

    Y.Map("sheet") 안 "snapshot" key에 fortune-sheet workbook JSON 저장됨.
    pycrdt로 디코드.
    """
    if not yjs_state:
        return None
    try:
        from pycrdt import Doc, Map
        doc = Doc()
        doc.apply_update(yjs_state)
        # Y.Map("sheet") — SheetEditor에서 사용 (HocuspocusProvider name="sheet-{id}")
        sheet_map = doc.get("sheet", type=Map)
        snap = sheet_map.get("snapshot")
        if isinstance(snap, list):
            return snap
        return None
    except Exception:
        return None


def _sheet_snapshot_to_xlsx(snapshot: list[dict] | None, fallback_title: str) -> bytes:
    """fortune-sheet snapshot → openpyxl XLSX bytes."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    if not snapshot:
        ws = wb.create_sheet(title=(fallback_title or "Sheet1")[:31])
        ws["A1"] = fallback_title or "데이터 없음"
        ws["A2"] = "원본은 본 시스템 yjs_state에 있고 외부 변환이 불완전합니다."
    else:
        for sheet_data in snapshot:
            ws_name = (sheet_data.get("name") or "Sheet")[:31]
            # 동일 시트명 충돌 회피
            counter = 0
            unique = ws_name
            while unique in wb.sheetnames:
                counter += 1
                unique = f"{ws_name[:28]}_{counter}"
            ws = wb.create_sheet(title=unique)
            cells = sheet_data.get("celldata", [])
            for c in cells:
                r = c.get("r", 0)
                col = c.get("c", 0)
                v = c.get("v", {})
                val = v.get("v") if isinstance(v, dict) else v
                if val is None:
                    continue
                cell = ws.cell(row=r + 1, column=col + 1, value=val)
                # 기본 서식 (굵게, 폰트 색상은 fortune-sheet의 ct/bl/fc 참조)
                if isinstance(v, dict):
                    try:
                        from openpyxl.styles import Font
                        bold = bool(v.get("bl"))
                        color = v.get("fc")
                        if bold or color:
                            cell.font = Font(
                                bold=bold,
                                color=(color.lstrip("#") if isinstance(color, str) else None),
                            )
                    except Exception:
                        pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _survey_to_csv(
    sv: Survey,
    questions: list[SurveyQuestion],
    responses: list[tuple[SurveyResponse, list[SurveyAnswer]]],
) -> str:
    """설문 → CSV 문자열.

    헤더: respondent_id (또는 anonymous), submitted_at, [질문1], [질문2], ...
    응답 1행 = 1 row. multi_choice는 |로 join.
    """
    import csv as _csv
    out = io.StringIO()
    # UTF-8 BOM (Excel 호환)
    out.write("﻿")

    headers = ["respondent_id", "submitted_at"]
    for q in questions:
        headers.append(f"{q.order + 1}. {q.question_text[:60]}")

    writer = _csv.writer(out)
    writer.writerow(headers)

    for resp, answers in responses:
        row = [
            "익명" if sv.is_anonymous else (resp.respondent_id or ""),
            resp.submitted_at.isoformat() if resp.submitted_at else "",
        ]
        ans_by_qid: dict[int, str] = {}
        for a in answers:
            if a.text_value:
                value = a.text_value
            elif a.choice_values:
                value = " | ".join(str(v) for v in a.choice_values)
            elif a.rating_value is not None:
                value = str(a.rating_value)
            else:
                value = ""
            ans_by_qid[a.question_id] = value
        for q in questions:
            row.append(ans_by_qid.get(q.id, ""))
        writer.writerow(row)
    return out.getvalue()


def _deck_to_html(p: ClassroomPresentation, slides: list[ClassroomSlide]) -> str:
    """프리젠테이션 → HTML 슬라이드 시리즈.

    각 슬라이드를 한 section으로. 인쇄 시 page break.
    캔버스 SVG 변환은 손실 큼 → plain_text + 제목만.
    """
    import html as _html
    title = _html.escape((p.title or "제목 없는 프리젠테이션").strip())
    sections = []
    for i, sl in enumerate(slides, start=1):
        sl_title = _html.escape((sl.title or "").strip())
        sl_text = (sl.plain_text or "").strip()
        body_paragraphs = "\n".join(
            f"<p>{_html.escape(line) if line.strip() else '&nbsp;'}</p>"
            for line in sl_text.split("\n")
        ) if sl_text else "<p><em>본문 비어 있음</em></p>"
        sections.append(f"""
  <section class="slide">
    <div class="slide-number">{i} / {len(slides)}</div>
    {f'<h2>{sl_title}</h2>' if sl_title else ''}
    {body_paragraphs}
  </section>""")

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <title>{title}</title>
  <style>
    body {{ font-family: -apple-system, "Segoe UI", "Malgun Gothic", sans-serif; max-width: 900px; margin: 2em auto; padding: 0 1em; color: #222; line-height: 1.6; }}
    h1 {{ border-bottom: 3px solid #673ab7; padding-bottom: 0.4em; }}
    .slide {{ border: 1px solid #ddd; border-radius: 6px; padding: 1.5em; margin: 1.2em 0; background: #fafafa; page-break-after: always; }}
    .slide-number {{ color: #888; font-size: 0.85em; margin-bottom: 0.5em; }}
    .slide h2 {{ color: #673ab7; margin-top: 0; }}
    .note {{ background: #fff8e1; border-left: 3px solid #f59e0b; padding: 0.8em; font-size: 0.9em; margin-top: 2em; }}
    @media print {{ .slide {{ page-break-after: always; }} }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  <p>슬라이드 {len(slides)}장 · 백업: {datetime.now(timezone.utc).isoformat()}</p>
  {''.join(sections)}
  <div class="note">
    각 슬라이드 본문 텍스트만 추출했습니다 (캔버스 도형·이미지는 손실).
    원본은 본 시스템에서 ZIP 복원 시 완전히 살아납니다.
  </div>
</body>
</html>
"""


def _prosemirror_xml_to_html(node) -> str:
    """pycrdt XmlElement/XmlText → HTML 재귀 변환.

    TipTap의 ProseMirror schema(paragraph/heading/list/table 등) 핵심 매핑.
    미지원 node는 children content만 보존.
    """
    import html as _h
    try:
        from pycrdt import XmlElement, XmlText
    except Exception:
        return ""

    if isinstance(node, XmlText):
        text = _h.escape(str(node))
        attrs = getattr(node, "attributes", None) or {}
        result = text
        if attrs.get("strong") or attrs.get("bold"):
            result = f"<strong>{result}</strong>"
        if attrs.get("em") or attrs.get("italic"):
            result = f"<em>{result}</em>"
        if attrs.get("code"):
            result = f"<code>{result}</code>"
        if attrs.get("strike"):
            result = f"<s>{result}</s>"
        link = attrs.get("link")
        if link:
            href = link.get("href", "#") if isinstance(link, dict) else str(link)
            result = f'<a href="{_h.escape(href)}">{result}</a>'
        return result

    if isinstance(node, XmlElement):
        tag = (getattr(node, "tag", "") or "").strip()
        attrs = getattr(node, "attributes", None) or {}
        children_html = "".join(_prosemirror_xml_to_html(c) for c in node.children)
        if tag == "paragraph":
            return f"<p>{children_html or '&nbsp;'}</p>"
        if tag == "heading":
            try:
                level = max(1, min(6, int(attrs.get("level", 1))))
            except (TypeError, ValueError):
                level = 1
            return f"<h{level}>{children_html}</h{level}>"
        if tag == "bulletList":
            return f"<ul>{children_html}</ul>"
        if tag == "orderedList":
            return f"<ol>{children_html}</ol>"
        if tag == "listItem":
            return f"<li>{children_html}</li>"
        if tag == "blockquote":
            return f"<blockquote>{children_html}</blockquote>"
        if tag == "codeBlock":
            return f"<pre><code>{children_html}</code></pre>"
        if tag == "horizontalRule":
            return "<hr>"
        if tag == "hardBreak":
            return "<br>"
        if tag == "table":
            return f"<table>{children_html}</table>"
        if tag == "tableRow":
            return f"<tr>{children_html}</tr>"
        if tag == "tableHeader":
            return f"<th>{children_html}</th>"
        if tag == "tableCell":
            return f"<td>{children_html}</td>"
        if tag == "image":
            src = _h.escape(str(attrs.get("src", "")))
            alt = _h.escape(str(attrs.get("alt", "")))
            return f'<img src="{src}" alt="{alt}">'
        # unknown — children만 보존
        return children_html

    return ""


def _try_decode_doc_html(yjs_state: bytes | None) -> str | None:
    """yjs_state → ProseMirror HTML (성공 시 본문 HTML만 반환)."""
    if not yjs_state:
        return None
    try:
        from pycrdt import Doc, XmlFragment
        doc = Doc()
        doc.apply_update(yjs_state)
        # TipTap Collaboration 기본 fragment 이름 = "default"
        frag = doc.get("default", type=XmlFragment)
        body = "".join(_prosemirror_xml_to_html(c) for c in frag.children)
        return body.strip() or None
    except Exception:
        return None


def _doc_to_html(d: ClassroomDocument) -> str:
    """문서 → HTML.

    Yjs CRDT(TipTap) → ProseMirror HTML 정밀 변환 시도. 실패 시 plain_text fallback.
    서식(굵게/기울임/목록/제목/표) 보존. 미지원 node는 텍스트만.
    """
    import html as _html
    title = (d.title or "제목 없는 문서").strip()
    title_esc = _html.escape(title)

    body_html = _try_decode_doc_html(d.yjs_state)
    fallback_used = body_html is None
    if fallback_used:
        body_text = (d.plain_text or "").strip()
        if body_text:
            body_html = "\n".join(
                f"<p>{_html.escape(line) if line.strip() else '&nbsp;'}</p>"
                for line in body_text.split("\n")
            )
        else:
            body_html = "<p><em>본문 없음</em></p>"

    note = (
        "본문이 평문화돼 있습니다 (Yjs 디코드 실패). 원본은 ZIP 복원 시 살아납니다."
        if fallback_used
        else "ProseMirror 구조 그대로 HTML 변환 — 굵게·목록·표 등 서식 보존."
    )

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <title>{title_esc}</title>
  <style>
    body {{ font-family: -apple-system, "Segoe UI", "Malgun Gothic", sans-serif; max-width: 800px; margin: 2em auto; padding: 0 1em; color: #222; line-height: 1.7; }}
    h1 {{ border-bottom: 2px solid #673ab7; padding-bottom: 0.3em; }}
    h2 {{ color: #555; border-bottom: 1px solid #eee; padding-bottom: 0.2em; }}
    .meta {{ color: #888; font-size: 0.85em; margin-bottom: 1em; }}
    blockquote {{ border-left: 4px solid #ddd; margin: 1em 0; padding: 0.3em 1em; color: #555; }}
    pre {{ background: #f4f4f4; padding: 0.8em; border-radius: 4px; overflow-x: auto; }}
    code {{ background: #f4f4f4; padding: 0.1em 0.3em; border-radius: 3px; font-size: 0.9em; }}
    table {{ border-collapse: collapse; margin: 1em 0; width: 100%; }}
    th, td {{ border: 1px solid #ccc; padding: 0.4em 0.6em; text-align: left; }}
    th {{ background: #f0f0f0; }}
    img {{ max-width: 100%; height: auto; }}
    .note {{ background: #fff8e1; border-left: 3px solid #f59e0b; padding: 0.8em; font-size: 0.9em; margin-top: 2em; }}
  </style>
</head>
<body>
  <h1>{title_esc}</h1>
  <div class="meta">백업 시각: {datetime.now(timezone.utc).isoformat()}</div>
  {body_html}
  <div class="note">{note}</div>
</body>
</html>
"""


async def _build_zip(db: AsyncSession, user: User) -> bytes:
    """본인 자료 모두 수집 → ZIP bytes 반환."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        # 폴더 트리
        folders = (await db.execute(
            select(Folder).where(
                Folder.owner_id == user.id,
                Folder.deleted_at.is_(None),
            ).order_by(Folder.sort_order)
        )).scalars().all()
        z.writestr(
            "folders.json",
            json.dumps([_serialize_folder(f) for f in folders], ensure_ascii=False, indent=2),
        )

        # docs — JSON (원본 + base64) + HTML (사람-읽기, 평문 기반)
        docs = (await db.execute(
            select(ClassroomDocument).where(
                ClassroomDocument.owner_id == user.id,
                ClassroomDocument.deleted_at.is_(None),
            )
        )).scalars().all()
        for d in docs:
            data = _serialize_doc(d)
            z.writestr(
                f"docs/{d.id}_{_safe_name(d.title)}.json",
                json.dumps(data, ensure_ascii=False, indent=2),
            )
            # pycrdt 디코드 + HTML 생성 (CPU bound) → to_thread
            try:
                html_text = await asyncio.to_thread(_doc_to_html, d)
                z.writestr(
                    f"docs/{d.id}_{_safe_name(d.title)}.html",
                    html_text,
                )
            except Exception:
                pass

        # sheets — JSON + XLSX (fortune-sheet snapshot 디코드)
        sheets = (await db.execute(
            select(ClassroomSheet).where(
                ClassroomSheet.owner_id == user.id,
                ClassroomSheet.deleted_at.is_(None),
            )
        )).scalars().all()
        for s in sheets:
            data = _serialize_sheet(s)
            z.writestr(
                f"sheets/{s.id}_{_safe_name(s.title)}.json",
                json.dumps(data, ensure_ascii=False, indent=2),
            )
            # pycrdt 디코드 + openpyxl Workbook.save 모두 CPU+IO 무거움 → to_thread
            try:
                snap = await asyncio.to_thread(
                    _extract_sheet_snapshot, s.yjs_state,
                )
                xlsx_bytes = await asyncio.to_thread(
                    _sheet_snapshot_to_xlsx, snap, s.title or "Sheet",
                )
                z.writestr(
                    f"sheets/{s.id}_{_safe_name(s.title)}.xlsx",
                    xlsx_bytes,
                )
            except Exception:
                pass

        # decks + slides
        decks = (await db.execute(
            select(ClassroomPresentation).where(
                ClassroomPresentation.owner_id == user.id,
                ClassroomPresentation.deleted_at.is_(None),
            )
        )).scalars().all()
        for p in decks:
            data = _serialize_deck(p)
            slides = (await db.execute(
                select(ClassroomSlide).where(
                    ClassroomSlide.presentation_id == p.id,
                ).order_by(ClassroomSlide.order)
            )).scalars().all()
            data["slides"] = [_serialize_slide(sl) for sl in slides]
            z.writestr(
                f"decks/{p.id}_{_safe_name(p.title)}.json",
                json.dumps(data, ensure_ascii=False, indent=2),
            )
            try:
                z.writestr(
                    f"decks/{p.id}_{_safe_name(p.title)}.html",
                    _deck_to_html(p, list(slides)),
                )
            except Exception:
                pass

        # surveys + questions + responses (CSV로도 export)
        surveys = (await db.execute(
            select(Survey).where(
                Survey.author_id == user.id,
                Survey.deleted_at.is_(None),
            )
        )).scalars().all()
        for sv in surveys:
            questions = (await db.execute(
                select(SurveyQuestion).where(
                    SurveyQuestion.survey_id == sv.id,
                ).order_by(SurveyQuestion.order)
            )).scalars().all()
            data = _serialize_survey(sv, questions)
            z.writestr(
                f"surveys/{sv.id}_{_safe_name(sv.title)}.json",
                json.dumps(data, ensure_ascii=False, indent=2),
            )
            # 응답 CSV (있을 때만)
            try:
                resps = (await db.execute(
                    select(SurveyResponse).where(
                        SurveyResponse.survey_id == sv.id,
                    ).order_by(SurveyResponse.submitted_at)
                )).scalars().all()
                resp_data: list[tuple] = []
                for resp in resps:
                    answers = (await db.execute(
                        select(SurveyAnswer).where(
                            SurveyAnswer.response_id == resp.id,
                        )
                    )).scalars().all()
                    resp_data.append((resp, list(answers)))
                if resp_data:
                    csv_text = _survey_to_csv(sv, questions, resp_data)
                    z.writestr(
                        f"surveys/{sv.id}_{_safe_name(sv.title)}_responses.csv",
                        csv_text.encode("utf-8"),
                    )
            except Exception:
                pass

        # hwps — file 그대로 + 메타
        hwps = (await db.execute(
            select(ClassroomHwp).where(
                ClassroomHwp.owner_id == user.id,
                ClassroomHwp.deleted_at.is_(None),
            )
        )).scalars().all()
        for h in hwps:
            z.writestr(
                f"hwps/{h.id}_{_safe_name(h.title)}.json",
                json.dumps(_serialize_hwp(h), ensure_ascii=False, indent=2),
            )
            if h.file_path:
                full = STORAGE_ROOT / h.file_path
                if full.exists():
                    fmt = h.file_format or "hwpx"
                    z.write(full, f"hwps/{h.id}_{_safe_name(h.title)}.{fmt}")

        # manifest
        manifest = {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "system": "general_school",
            "version": "1.0",
            "user": {
                "id": user.id,
                "email": user.email,
                "name": user.name,
                "role": user.role,
            },
            "counts": {
                "folders": len(folders),
                "docs": len(docs),
                "sheets": len(sheets),
                "decks": len(decks),
                "surveys": len(surveys),
                "hwps": len(hwps),
            },
            "note": (
                "본 ZIP은 일반학교 플랫폼(general_school) 백업입니다. "
                "각 자료 폴더에 두 가지 형식 포함:\n"
                " - JSON: 본 시스템 재import 용 (yjs_state binary base64 보존)\n"
                " - 사람-읽기: docs/*.html (평문), sheets/*.xlsx (Excel), "
                "surveys/*_responses.csv (응답), hwps/*.hwpx (원본)\n"
                "사람-읽기 형식은 어느 PC에서나 즉시 열림. JSON은 같은 시스템 복원용."
            ),
        }
        z.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    return buf.getvalue()


@router.post("/backup/download")
async def download_my_drive_backup(
    user: User = Depends(require_permission("drive.use")),
    db: AsyncSession = Depends(get_db),
):
    """본인 드라이브 전체를 ZIP으로 다운로드.

    학교 이동 시 사용. 자료가 많으면 응답 지연 가능 (background queue는 향후).
    """
    blob = await _build_zip(db, user)
    fname = f"drive-backup-{user.email}-{datetime.now().strftime('%Y%m%d')}.zip"

    await log_action(
        db, user, "drive.backup.download",
        target=f"user:{user.id}",
        detail=f"size={len(blob)}",
    )

    return StreamingResponse(
        io.BytesIO(blob),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Import (복원)
# ─────────────────────────────────────────────────────────────────────────────

# 안전 한도 — DoS 차단
MAX_IMPORT_ZIP_SIZE = 500 * 1024 * 1024  # 500MB
MAX_ITEMS_PER_TYPE = 2000  # 한 사용자 한 번에 자료 2000개 한도


async def _resolve_imported_folder(
    db: AsyncSession, user: User, src_folder: dict[str, Any],
    folder_id_map: dict[int, int],
) -> int | None:
    """ZIP 안 폴더 한 개를 사용자 드라이브에 매핑.

    - 자동 폴더(is_system_locked): 기존 폴더 찾아서 그대로 사용 (auto_kind + source 매칭).
      없으면 새로 생성 (단 잠금 유지).
    - 수동 폴더: 새 생성 (이름 중복 허용).
    - 결과는 folder_id_map에 (옛 id → 새 id) 저장.
    """
    old_id = src_folder.get("id")
    if old_id is None:
        return None

    # 자동 폴더 — 멱등 매칭
    if src_folder.get("is_system_locked") and src_folder.get("auto_kind"):
        q = select(Folder).where(
            Folder.owner_id == user.id,
            Folder.auto_kind == src_folder["auto_kind"],
        )
        if src_folder.get("semester_id") is not None:
            q = q.where(Folder.semester_id == src_folder["semester_id"])
        else:
            q = q.where(Folder.semester_id.is_(None))
        if src_folder.get("source_kind") is not None:
            q = q.where(Folder.source_kind == src_folder["source_kind"])
        if src_folder.get("source_id") is not None:
            q = q.where(Folder.source_id == src_folder["source_id"])
        existing = (await db.execute(q.limit(1))).scalar_one_or_none()
        if existing:
            folder_id_map[old_id] = existing.id
            return existing.id

    # 새 폴더 — 부모 매핑
    parent_old = src_folder.get("parent_id")
    parent_new = folder_id_map.get(parent_old) if parent_old else None

    # 다음 sort_order
    if parent_new is None:
        max_order = (await db.execute(
            select(Folder.sort_order).where(
                Folder.owner_id == user.id, Folder.parent_id.is_(None),
                Folder.deleted_at.is_(None),
            ).order_by(Folder.sort_order.desc()).limit(1)
        )).scalar_one_or_none() or 0
    else:
        max_order = (await db.execute(
            select(Folder.sort_order).where(
                Folder.parent_id == parent_new, Folder.deleted_at.is_(None),
            ).order_by(Folder.sort_order.desc()).limit(1)
        )).scalar_one_or_none() or 0

    f = Folder(
        owner_id=user.id,
        parent_id=parent_new,
        name=(src_folder.get("name") or "복원된 폴더")[:255],
        auto_kind=src_folder.get("auto_kind"),
        semester_id=src_folder.get("semester_id"),
        source_kind=src_folder.get("source_kind"),
        source_id=src_folder.get("source_id"),
        sort_order=int(max_order) + 1,
        is_system_locked=bool(src_folder.get("is_system_locked")),
    )
    db.add(f)
    await db.flush()
    folder_id_map[old_id] = f.id
    return f.id


@router.post("/backup/import")
async def import_drive_backup(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_permission("drive.use")),
    db: AsyncSession = Depends(get_db),
):
    """본인 드라이브에 ZIP 백업 복원.

    원칙:
      - 본인 드라이브에만 — 다른 사용자 자료 복원 불가
      - 자동 폴더는 기존과 매칭 (중복 생성 X). 수동 폴더는 새로 생성.
      - 자료는 모두 새 id로 생성 — 기존 자료 안 건드림
      - HWP file은 새 storage 경로로 복사
      - quota check_quota 통과 + consume_quota
      - 한 번에 500MB ZIP / 자료 type당 2000개 한도 (DoS 차단)
    """
    # 1) 확장자·크기 검증 (POLICY_BACKUP: .zip, 2GB)
    content = await validate_upload(file, POLICY_BACKUP)
    if len(content) > MAX_IMPORT_ZIP_SIZE:
        raise HTTPException(413, f"ZIP이 너무 큽니다 (최대 {MAX_IMPORT_ZIP_SIZE // 1024 // 1024}MB)")

    # 2) ZIP 파싱
    try:
        zf = zipfile.ZipFile(io.BytesIO(content), "r")
    except zipfile.BadZipFile:
        raise HTTPException(400, "유효한 ZIP 파일이 아닙니다")

    names = set(zf.namelist())
    if "manifest.json" not in names:
        raise HTTPException(400, "manifest.json 없음 — 본 시스템 백업 ZIP 아닐 수 있음")

    try:
        manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
    except Exception:
        raise HTTPException(400, "manifest.json 파싱 실패")
    if manifest.get("system") != "general_school":
        raise HTTPException(400, f"호환 안 되는 system: {manifest.get('system')}")

    # 3) Quota 추정 — 자료 storage_bytes 합계
    estimated_bytes = 0
    items_to_import: dict[str, list[dict]] = {
        "docs": [], "sheets": [], "decks": [], "surveys": [], "hwps": [],
    }
    for n in sorted(names):
        for t in items_to_import.keys():
            if n.startswith(f"{t}/") and n.endswith(".json"):
                try:
                    data = json.loads(zf.read(n).decode("utf-8"))
                    items_to_import[t].append(data)
                    estimated_bytes += int(data.get("storage_bytes") or 0)
                except Exception:
                    pass
        if len(items_to_import["docs"]) > MAX_ITEMS_PER_TYPE:
            raise HTTPException(413, f"docs 자료가 {MAX_ITEMS_PER_TYPE}개 초과")

    if estimated_bytes > 0:
        check_quota(user, estimated_bytes)

    # 4) 폴더 import (folders.json)
    folder_id_map: dict[int, int] = {}
    if "folders.json" in names:
        try:
            folders_data = json.loads(zf.read("folders.json").decode("utf-8"))
        except Exception:
            folders_data = []
        # parent_id 의존 순서 위해 두 패스
        # 1차: 모든 폴더를 임시 등록 (parent 매핑 후일에 처리)
        remaining = list(folders_data)
        # parent_id 가 None 또는 이미 매핑된 폴더부터 처리. 최대 64 depth.
        for _ in range(64):
            if not remaining:
                break
            progress = False
            still: list[dict] = []
            for sf in remaining:
                parent_old = sf.get("parent_id")
                if parent_old is None or parent_old in folder_id_map:
                    await _resolve_imported_folder(db, user, sf, folder_id_map)
                    progress = True
                else:
                    still.append(sf)
            remaining = still
            if not progress:
                break  # cycle 또는 invalid

    # 5) 자료 import
    imported_counts = {"folders": len(folder_id_map), "docs": 0, "sheets": 0, "decks": 0, "surveys": 0, "hwps": 0}

    # docs
    for d in items_to_import["docs"]:
        new_obj = ClassroomDocument(
            owner_id=user.id,
            course_id=None,  # 강좌 상관 X (본인 드라이브 복원)
            title=(d.get("title") or "복원된 문서")[:255],
            plain_text=d.get("plain_text"),
            yjs_state=base64.b64decode(d["yjs_state_base64"]) if d.get("yjs_state_base64") else None,
            access_mode=d.get("access_mode") or "specific_users",
            storage_bytes=int(d.get("storage_bytes") or 0),
            folder_id=folder_id_map.get(d.get("folder_id")) if d.get("folder_id") else None,
        )
        db.add(new_obj)
        imported_counts["docs"] += 1

    # sheets
    for s in items_to_import["sheets"]:
        new_obj = ClassroomSheet(
            owner_id=user.id,
            course_id=None,
            title=(s.get("title") or "복원된 시트")[:255],
            yjs_state=base64.b64decode(s["yjs_state_base64"]) if s.get("yjs_state_base64") else None,
            access_mode=s.get("access_mode") or "specific_users",
            settings=s.get("settings"),
            storage_bytes=int(s.get("storage_bytes") or 0),
            folder_id=folder_id_map.get(s.get("folder_id")) if s.get("folder_id") else None,
        )
        db.add(new_obj)
        imported_counts["sheets"] += 1

    # decks + slides
    for p in items_to_import["decks"]:
        new_obj = ClassroomPresentation(
            owner_id=user.id,
            course_id=None,
            title=(p.get("title") or "복원된 프리젠테이션")[:255],
            yjs_state=base64.b64decode(p["yjs_state_base64"]) if p.get("yjs_state_base64") else None,
            access_mode=p.get("access_mode") or "specific_users",
            settings=p.get("settings"),
            storage_bytes=int(p.get("storage_bytes") or 0),
            folder_id=folder_id_map.get(p.get("folder_id")) if p.get("folder_id") else None,
        )
        db.add(new_obj)
        await db.flush()
        # slides
        for sl in (p.get("slides") or []):
            db.add(ClassroomSlide(
                presentation_id=new_obj.id,
                order=int(sl.get("order") or 0),
                title=sl.get("title"),
                plain_text=sl.get("plain_text"),
                settings=sl.get("settings"),
            ))
        imported_counts["decks"] += 1

    # surveys + questions (응답은 복원 X — 새 학교에선 새 응답)
    for sv in items_to_import["surveys"]:
        new_obj = Survey(
            author_id=user.id,
            course_id=None,
            title=(sv.get("title") or "복원된 설문")[:255],
            description=sv.get("description"),
            status="draft",  # 안전: 복원 시 draft로 (active 즉시 응답 받지 않게)
            is_anonymous=bool(sv.get("is_anonymous")),
            access_mode=sv.get("access_mode") or "course_members",
            storage_bytes=int(sv.get("storage_bytes") or 0),
            folder_id=folder_id_map.get(sv.get("folder_id")) if sv.get("folder_id") else None,
        )
        db.add(new_obj)
        await db.flush()
        for q in (sv.get("questions") or []):
            db.add(SurveyQuestion(
                survey_id=new_obj.id,
                order=int(q.get("order") or 0),
                question_type=q.get("question_type") or q.get("type") or "short_text",
                question_text=q.get("question_text") or "질문",
                is_required=bool(q.get("is_required")),
                options=q.get("options"),
            ))
        imported_counts["surveys"] += 1

    # hwps + file 복사
    STORAGE_HWPS = STORAGE_ROOT / "hwps"
    for h in items_to_import["hwps"]:
        new_obj = ClassroomHwp(
            owner_id=user.id,
            course_id=None,
            title=(h.get("title") or "복원된 HWP")[:255],
            access_mode=h.get("access_mode") or "specific_users",
            file_format=h.get("file_format"),
            storage_bytes=int(h.get("storage_bytes") or 0),
            folder_id=folder_id_map.get(h.get("folder_id")) if h.get("folder_id") else None,
        )
        db.add(new_obj)
        await db.flush()
        # ZIP 안에 hwp 파일 찾기 — hwps/{old_id}_*.{hwp|hwpx}
        old_id = h.get("id")
        fmt = h.get("file_format") or "hwpx"
        # 파일명 prefix 찾기 (id_제목.fmt)
        for n in names:
            if n.startswith(f"hwps/{old_id}_") and (n.endswith(".hwp") or n.endswith(".hwpx")):
                try:
                    file_data = zf.read(n)
                    token = secrets.token_urlsafe(8)
                    new_dir = STORAGE_HWPS / str(new_obj.id)
                    await ensure_dir_async(new_dir)
                    new_fname = f"{token}.{fmt}"
                    await write_bytes_async(new_dir / new_fname, file_data)
                    new_obj.file_path = f"hwps/{new_obj.id}/{new_fname}"
                except Exception:
                    pass
                break
        imported_counts["hwps"] += 1

    await db.flush()

    # 6) quota consume
    if estimated_bytes > 0:
        await consume_quota(db, user, estimated_bytes)

    await log_action(
        db, user, "drive.backup.import",
        target=f"user:{user.id}",
        detail=(
            f"folders={imported_counts['folders']} "
            f"docs={imported_counts['docs']} sheets={imported_counts['sheets']} "
            f"decks={imported_counts['decks']} surveys={imported_counts['surveys']} "
            f"hwps={imported_counts['hwps']} bytes={estimated_bytes}"
        ),
        request=request,
    )

    return {
        "ok": True,
        "imported": imported_counts,
        "consumed_bytes": estimated_bytes,
        "source_user": manifest.get("user"),
        "note": (
            "복원 완료. 설문지는 draft 상태로 들어왔습니다 — 게시하려면 설정에서 "
            "변경하세요. HWP 파일은 새 storage 경로에 복사됐습니다."
        ),
    }



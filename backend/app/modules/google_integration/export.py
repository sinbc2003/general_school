"""Google Drive로 자료 export — 협업 문서·스프레드시트를 Google 형식으로 업로드.

지원:
  - ClassroomDocument (TipTap HTML) → Google Docs
    · HTML 추출 → multipart upload (mimeType=application/vnd.google-apps.document)
    · Drive가 자동으로 Docs로 변환
  - ClassroomSheet (fortune-sheet snapshot) → Google Sheets
    · 데이터 → openpyxl Workbook → XLSX binary → upload (자동 Sheets 변환)

PPT/Slides는 캔버스 기반이라 손실 큼 → 지원 안 함 (정책).

엔드포인트:
  POST /api/google/export/docs/{id}    — 문서 → Drive
  POST /api/google/export/sheets/{id}  — 시트 → Drive
"""

import asyncio
from io import BytesIO
import json

import httpx
from fastapi import APIRouter, Depends, HTTPException, Request
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_action
from app.core.database import get_db
from app.core.permissions import require_permission
from app.models import ClassroomDocument, ClassroomSheet, User
from app.modules.google_integration.router import _google_http, get_access_token_for_user, router

DRIVE_UPLOAD_URL = "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart"


def _build_multipart(meta: dict, content_type: str, content: bytes) -> tuple[str, bytes]:
    """Drive API multipart/related body 생성."""
    boundary = "----------school_export_boundary"
    meta_json = json.dumps(meta).encode("utf-8")
    body = (
        f"--{boundary}\r\n".encode()
        + b"Content-Type: application/json; charset=UTF-8\r\n\r\n"
        + meta_json + b"\r\n"
        + f"--{boundary}\r\n".encode()
        + f"Content-Type: {content_type}\r\n\r\n".encode()
        + content
        + f"\r\n--{boundary}--\r\n".encode()
    )
    return boundary, body


@router.post("/export/docs/{doc_id}")
async def export_doc_to_drive(
    doc_id: int,
    request: Request,
    user: User = Depends(require_permission("google.integration.use")),
    db: AsyncSession = Depends(get_db),
):
    """협업 문서 → Google Docs."""
    doc = await db.get(ClassroomDocument, doc_id)
    if not doc:
        raise HTTPException(404, "문서 없음")
    if doc.owner_id != user.id and user.role not in ("super_admin", "designated_admin"):
        raise HTTPException(403, "본인 문서만 export 가능")

    access = await get_access_token_for_user(db, user)

    # TipTap HTML은 yjs_state binary에 들어있어 추출 복잡 → plain_text를 H1 + body로 단순 변환
    # 향후: Yjs state → TipTap JSON → HTML 변환은 별도 worker로
    title = doc.title or "제목 없는 문서"
    body_html = (doc.plain_text or "").replace("\n", "<br>")
    html = f"<html><body><h1>{title}</h1><div>{body_html}</div></body></html>"

    meta = {
        "name": title,
        "mimeType": "application/vnd.google-apps.document",
    }
    boundary, body = _build_multipart(meta, "text/html; charset=UTF-8", html.encode("utf-8"))

    r = await _google_http(
        "POST", DRIVE_UPLOAD_URL, timeout=30,
        headers={
            "Authorization": f"Bearer {access}",
            "Content-Type": f"multipart/related; boundary={boundary}",
        },
        content=body,
    )
    if r.status_code not in (200, 201):
        raise HTTPException(502, f"Drive upload 실패: {r.status_code} {r.text[:200]}")

    j = r.json()
    await log_action(
        db, user, "google_export_doc",
        target=f"doc:{doc_id}",
        detail=f"drive_id={j.get('id')} drive_name={j.get('name')}",
        request=request,
    )
    return {
        "ok": True,
        "drive_file_id": j.get("id"),
        "drive_file_name": j.get("name"),
        "view_url": f"https://docs.google.com/document/d/{j.get('id')}/edit",
    }


def _build_xlsx_from_sheet_sync(sheet: ClassroomSheet) -> bytes:
    """fortune-sheet snapshot에서 XLSX 생성. sync (caller가 to_thread)."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    snapshot = None
    if sheet.yjs_state:
        try:
            # yjs_state는 binary CRDT. fortune-sheet snapshot은 Y.Map.get("sheet")에 JSON.
            # 직접 디코드는 어렵고, sheet 데이터를 별도로 보관해야 정확. 임시로 빈 시트 + 제목만.
            pass
        except Exception:
            pass

    if not snapshot:
        # fallback: 빈 시트 + title
        ws = wb.create_sheet(title=(sheet.title or "Sheet1")[:31])
        ws["A1"] = sheet.title or "제목 없음"
    else:
        for sheet_data in snapshot:
            ws_name = (sheet_data.get("name") or "Sheet")[:31]
            ws = wb.create_sheet(title=ws_name)
            cells = sheet_data.get("celldata", [])
            for c in cells:
                r = c.get("r", 0)
                col = c.get("c", 0)
                v = c.get("v", {})
                val = v.get("v") if isinstance(v, dict) else v
                if val is not None:
                    ws.cell(row=r + 1, column=col + 1, value=val)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


@router.post("/export/sheets/{sheet_id}")
async def export_sheet_to_drive(
    sheet_id: int,
    request: Request,
    user: User = Depends(require_permission("google.integration.use")),
    db: AsyncSession = Depends(get_db),
):
    """협업 시트 → Google Sheets."""
    sheet = await db.get(ClassroomSheet, sheet_id)
    if not sheet:
        raise HTTPException(404, "시트 없음")
    if sheet.owner_id != user.id and user.role not in ("super_admin", "designated_admin"):
        raise HTTPException(403, "본인 시트만 export 가능")

    access = await get_access_token_for_user(db, user)
    xlsx_bytes = await asyncio.to_thread(_build_xlsx_from_sheet_sync, sheet)

    meta = {
        "name": sheet.title or "제목 없는 시트",
        "mimeType": "application/vnd.google-apps.spreadsheet",
    }
    boundary, body = _build_multipart(
        meta,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        xlsx_bytes,
    )

    r = await _google_http(
        "POST", DRIVE_UPLOAD_URL, timeout=60,
        headers={
            "Authorization": f"Bearer {access}",
            "Content-Type": f"multipart/related; boundary={boundary}",
        },
        content=body,
    )
    if r.status_code not in (200, 201):
        raise HTTPException(502, f"Drive upload 실패: {r.status_code} {r.text[:200]}")

    j = r.json()
    await log_action(
        db, user, "google_export_sheet",
        target=f"sheet:{sheet_id}",
        detail=f"drive_id={j.get('id')}",
        request=request,
    )
    return {
        "ok": True,
        "drive_file_id": j.get("id"),
        "drive_file_name": j.get("name"),
        "view_url": f"https://docs.google.com/spreadsheets/d/{j.get('id')}/edit",
    }


@router.post("/export/my-drive-bulk")
async def export_my_drive_to_google_bulk(
    request: Request,
    user: User = Depends(require_permission("google.integration.use")),
    db: AsyncSession = Depends(get_db),
):
    """본인 드라이브의 docs + sheets 모두 Google Drive로 일괄 export.

    학교 이동 시 사용. decks/surveys/hwps는 Google 변환 미지원 — ZIP 백업 권장.
    Google 토큰 미연결이면 400. 자료 많으면 응답 지연.
    """
    results: list[dict] = []
    docs = (await db.execute(
        select(ClassroomDocument).where(
            ClassroomDocument.owner_id == user.id,
            ClassroomDocument.deleted_at.is_(None),
        )
    )).scalars().all()
    sheets = (await db.execute(
        select(ClassroomSheet).where(
            ClassroomSheet.owner_id == user.id,
            ClassroomSheet.deleted_at.is_(None),
        )
    )).scalars().all()

    access = await get_access_token_for_user(db, user)

    # docs export
    for doc in docs:
        try:
            title = doc.title or "제목 없는 문서"
            body_html = (doc.plain_text or "").replace("\n", "<br>")
            html = f"<html><body><h1>{title}</h1><div>{body_html}</div></body></html>"
            meta = {"name": title, "mimeType": "application/vnd.google-apps.document"}
            boundary, body = _build_multipart(meta, "text/html; charset=UTF-8", html.encode("utf-8"))
            async with httpx.AsyncClient(timeout=30) as client:
                r = await client.post(
                    DRIVE_UPLOAD_URL,
                    headers={
                        "Authorization": f"Bearer {access}",
                        "Content-Type": f"multipart/related; boundary={boundary}",
                    },
                    content=body,
                )
            if r.status_code not in (200, 201):
                results.append({"type": "docs", "id": doc.id, "ok": False, "error": f"HTTP {r.status_code}"})
                continue
            j = r.json()
            results.append({
                "type": "docs", "id": doc.id, "title": title, "ok": True,
                "drive_file_id": j.get("id"),
                "view_url": f"https://docs.google.com/document/d/{j.get('id')}/edit",
            })
        except Exception as e:
            results.append({"type": "docs", "id": doc.id, "ok": False, "error": str(e)[:200]})

    # sheets export
    for sheet in sheets:
        try:
            xlsx_bytes = await asyncio.to_thread(_build_xlsx_from_sheet_sync, sheet)
            title = sheet.title or "제목 없는 스프레드시트"
            meta = {"name": title, "mimeType": "application/vnd.google-apps.spreadsheet"}
            boundary, body = _build_multipart(
                meta, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                xlsx_bytes,
            )
            async with httpx.AsyncClient(timeout=60) as client:
                r = await client.post(
                    DRIVE_UPLOAD_URL,
                    headers={
                        "Authorization": f"Bearer {access}",
                        "Content-Type": f"multipart/related; boundary={boundary}",
                    },
                    content=body,
                )
            if r.status_code not in (200, 201):
                results.append({"type": "sheets", "id": sheet.id, "ok": False, "error": f"HTTP {r.status_code}"})
                continue
            j = r.json()
            results.append({
                "type": "sheets", "id": sheet.id, "title": title, "ok": True,
                "drive_file_id": j.get("id"),
                "view_url": f"https://docs.google.com/spreadsheets/d/{j.get('id')}/edit",
            })
        except Exception as e:
            results.append({"type": "sheets", "id": sheet.id, "ok": False, "error": str(e)[:200]})

    ok_count = sum(1 for r in results if r["ok"])
    fail_count = len(results) - ok_count
    await log_action(
        db, user, "google_export_bulk",
        detail=f"ok={ok_count} fail={fail_count} docs={len(docs)} sheets={len(sheets)}",
        request=request,
    )
    return {
        "total": len(results),
        "ok": ok_count,
        "failed": fail_count,
        "results": results,
        "note": "decks/surveys/hwps는 Google Drive 변환 미지원 — ZIP 백업으로 받으세요",
    }

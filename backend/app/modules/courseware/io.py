"""문제은행 코스웨어 — JSONL/ZIP import + 결과 export (CSV/XLSX).

JSONL 형식 (한 줄 = 한 문제):
  {"type": "multiple_choice", "content": "1+1은?",
   "answer_data": {"grader_type": "choices", "correct": ["B"],
                   "choices": ["A. 1", "B. 2", "C. 3"]},
   "answer": "2", "difficulty": "easy", "subject": "수학", "tags": ["기초"]}

ZIP 형식 (이미지 포함):
  math.zip
   ├ problems.jsonl   # content 안에 ![](images/fig1.png)
   └ images/
      ├ fig1.png
      └ fig2.png
  → 백엔드가 풀어서 이미지를 storage/courseware/{token}.{ext}에 저장하고
    content 안의 'images/X.ext' 경로를 '/api/files/storage/courseware/Y.ext'로 치환.

router 객체는 router.py에서 공유. router.py 끝의 'from . import io'로 등록.
"""

from __future__ import annotations

import asyncio
import csv
import io as _io
import json
import re
import secrets
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from fastapi import Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_action
from app.core.database import get_db
from app.core.files import ensure_dir_async, write_bytes_async
from app.core.permissions import require_permission
from app.core.upload import POLICY_PROBLEMS_JSONL, POLICY_PROBLEMS_ZIP, validate_upload
from app.models import (
    Course, CourseProblemSet, CourseStudent, Problem,
    StudentProblemAttempt, User,
)
from app.modules.classroom.teachers import is_course_editor_or_admin
from app.modules.courseware.router import router
from app.modules.courseware.schemas import ProblemInline


# Storage paths — settings.STORAGE_ROOT 기반 (Phase 2-Q 통합).
from app.core.files import DEFAULT_STORAGE_ROOT
COURSEWARE_STORAGE_DIR = DEFAULT_STORAGE_ROOT / "courseware"
ALLOWED_IMAGE_EXTS = frozenset({".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"})
MAX_IMAGES_PER_ZIP = 500
MAX_ZIP_UNCOMPRESSED = 200 * 1024 * 1024  # 200MB (zip-bomb 방어)


# ─────────────────────────────────────────────────────────────────────────────
# 헬퍼
# ─────────────────────────────────────────────────────────────────────────────

MAX_JSONL_PROBLEMS = 5000  # 한 번의 import 최대 문제 수


def _parse_jsonl(raw: bytes) -> list[tuple[int, dict | str]]:
    """JSONL 파싱. 각 줄을 dict 또는 error string으로 반환.

    반환: [(line_no, parsed_dict | error_message), ...]
    빈 줄은 skip.
    """
    text = raw.decode("utf-8", errors="replace")
    out: list[tuple[int, dict | str]] = []
    for i, line in enumerate(text.splitlines(), start=1):
        s = line.strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except json.JSONDecodeError as e:
            out.append((i, f"JSON parse error: {e.msg} at col {e.colno}"))
            continue
        if not isinstance(obj, dict):
            out.append((i, "각 줄은 JSON object여야 합니다 (배열·문자열 X)"))
            continue
        out.append((i, obj))
    return out


def _validate_inline(obj: dict) -> tuple[ProblemInline | None, str | None]:
    """dict → ProblemInline 변환 + answer_data 무결성 1차 검증.

    실패 시 (None, error_message). 성공 시 (ProblemInline, None).
    """
    try:
        pi = ProblemInline(**obj)
    except Exception as e:  # Pydantic ValidationError 포함
        return (None, f"schema error: {e}")
    ad = pi.answer_data or {}
    grader = (ad.get("grader_type") or "").strip().lower()
    if not grader:
        return (None, "answer_data.grader_type 필수 (choices/exact/regex/numeric/essay/manual/llm)")
    if grader == "choices" and not ad.get("correct"):
        return (None, "choices grader는 answer_data.correct (정답 list) 필요")
    if grader == "exact" and ad.get("correct") is None:
        return (None, "exact grader는 answer_data.correct 필요")
    if grader == "regex" and not ad.get("pattern"):
        return (None, "regex grader는 answer_data.pattern 필요")
    if grader == "numeric" and ad.get("value") is None:
        return (None, "numeric grader는 answer_data.value 필요")
    return (pi, None)


# ─────────────────────────────────────────────────────────────────────────────
# JSONL import
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/courses/{cid}/problems/import-jsonl")
async def import_problems_jsonl(
    cid: int,
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(default=False, description="True면 검증만 (DB 변경 X)"),
    create_set: bool = Query(default=True, description="True면 import 성공 후 ProblemSet도 함께 생성 (status=draft)"),
    set_title: str | None = Query(default=None, description="create_set=True일 때 ProblemSet 제목"),
    user: User = Depends(require_permission("classroom.courseware.create")),
    db: AsyncSession = Depends(get_db),
):
    """JSONL 파일 한 줄당 한 문제 일괄 등록.

    응답:
      {
        "total": int,
        "valid": int,
        "errors": [{"line": N, "message": "..."}, ...],
        "created_problem_ids": [...],
        "problem_set_id": int | null,
      }
    """
    course = await db.get(Course, cid)
    if not course:
        raise HTTPException(404, "강좌 없음")
    if not await is_course_editor_or_admin(db, course, user):
        raise HTTPException(403, "강좌 교사·관리자만 가능")

    data = await validate_upload(file, POLICY_PROBLEMS_JSONL)
    parsed = _parse_jsonl(data)
    if len(parsed) > MAX_JSONL_PROBLEMS:
        raise HTTPException(
            400,
            f"한 번에 최대 {MAX_JSONL_PROBLEMS}문제까지 import 가능 "
            f"(현재 {len(parsed)}줄). 여러 파일로 나눠주세요.",
        )

    errors: list[dict] = []
    inlines: list[tuple[int, ProblemInline]] = []
    for line_no, item in parsed:
        if isinstance(item, str):
            errors.append({"line": line_no, "message": item})
            continue
        pi, err = _validate_inline(item)
        if err or pi is None:
            errors.append({"line": line_no, "message": err or "unknown error"})
            continue
        inlines.append((line_no, pi))

    result: dict = {
        "total": len(parsed),
        "valid": len(inlines),
        "errors": errors,
        "created_problem_ids": [],
        "problem_set_id": None,
    }

    if dry_run:
        return result
    if not inlines:
        raise HTTPException(400, "유효한 문제가 한 줄도 없습니다.")

    # Problem row 일괄 생성
    new_objs: list[Problem] = []
    for _line_no, pi in inlines:
        obj = Problem(
            department="math",
            subject=pi.subject or "",
            difficulty=pi.difficulty,
            question_type=pi.type,
            content=pi.content,
            solution=pi.solution,
            answer=pi.answer,
            answer_data=pi.answer_data,
            tags=pi.tags,
            is_visible=True,
            review_status="pending",
            created_by_id=user.id,
        )
        db.add(obj)
        new_objs.append(obj)
    await db.flush()
    created_ids = [o.id for o in new_objs]
    result["created_problem_ids"] = created_ids

    # ProblemSet 함께 생성 (옵션)
    if create_set:
        title = (set_title or "").strip() or f"JSONL Import {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ps = CourseProblemSet(
            course_id=cid,
            title=title,
            description=f"JSONL import — {len(created_ids)}문제",
            problem_ids=created_ids,
            status="draft",
            created_by=user.id,
        )
        db.add(ps)
        await db.flush()
        result["problem_set_id"] = ps.id

    await log_action(
        db, user, "courseware.problems.import_jsonl",
        target=f"course:{cid} count:{len(created_ids)}",
        detail=f"errors:{len(errors)} set:{result['problem_set_id']}",
        request=request,
    )
    return result


# ─────────────────────────────────────────────────────────────────────────────
# ZIP import (이미지 포함)
# ─────────────────────────────────────────────────────────────────────────────


def _safe_zip_extract_paths(zf: zipfile.ZipFile) -> list[zipfile.ZipInfo]:
    """ZIP entry들을 path traversal 차단하면서 정상화. 반환 = 통과한 entry.

    - 절대 경로 차단
    - ``..`` 포함 차단
    - 백슬래시(Windows zip) → forward slash 정규화
    - 디렉터리(끝이 '/') skip
    - 압축 해제 후 누적 크기 200MB 초과 시 차단 (zip-bomb)
    """
    out: list[zipfile.ZipInfo] = []
    total = 0
    for info in zf.infolist():
        name = info.filename.replace("\\", "/")
        if not name or name.endswith("/"):
            continue
        if name.startswith("/") or ".." in name.split("/"):
            raise HTTPException(400, f"안전하지 않은 경로: {info.filename}")
        info.filename = name
        total += info.file_size
        if total > MAX_ZIP_UNCOMPRESSED:
            raise HTTPException(
                400, f"ZIP 압축 해제 크기 초과: {total} > {MAX_ZIP_UNCOMPRESSED} bytes",
            )
        out.append(info)
    return out


def _find_jsonl_entry(entries: list[zipfile.ZipInfo]) -> zipfile.ZipInfo | None:
    """problems.jsonl, problems.json, *.jsonl 순으로 찾음."""
    by_name: dict[str, zipfile.ZipInfo] = {e.filename.lower(): e for e in entries}
    for cand in ("problems.jsonl", "problems.json"):
        if cand in by_name:
            return by_name[cand]
    for e in entries:
        if e.filename.lower().endswith(".jsonl"):
            return e
    return None


def _rewrite_image_paths(content: str, image_map: dict[str, str]) -> str:
    """content 안의 마크다운 이미지 경로를 새 URL로 치환.

    매칭 패턴 (대소문자 무관):
      ![alt](images/foo.png)
      ![alt](./images/foo.png)
      ![alt](/images/foo.png)
      ![alt](IMAGES/foo.png)
    """
    if not image_map:
        return content
    result = content
    for orig_key, new_url in image_map.items():
        # 원본 키는 'images/foo.png' 형식, 선두 ./ 또는 / 허용
        escaped = re.escape(orig_key)
        pattern = re.compile(
            rf'!\[([^\]]*)\]\(\s*\.?/?{escaped}\s*\)',
            flags=re.IGNORECASE,
        )
        result = pattern.sub(lambda m: f"![{m.group(1)}]({new_url})", result)
    return result


@router.post("/courses/{cid}/problems/import-zip")
async def import_problems_zip(
    cid: int,
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(default=False),
    create_set: bool = Query(default=True),
    set_title: str | None = Query(default=None),
    user: User = Depends(require_permission("classroom.courseware.create")),
    db: AsyncSession = Depends(get_db),
):
    """ZIP 패키지 import — problems.jsonl + images/ 폴더 묶음.

    이미지는 storage/courseware/{nanoid}.{ext}로 저장, content 안의
    `![](images/foo.png)` 패턴을 새 URL로 자동 치환.

    응답: JSONL import와 동일 + ``imported_images`` (이미지 개수).
    """
    course = await db.get(Course, cid)
    if not course:
        raise HTTPException(404, "강좌 없음")
    if not await is_course_editor_or_admin(db, course, user):
        raise HTTPException(403, "강좌 교사·관리자만 가능")

    data = await validate_upload(file, POLICY_PROBLEMS_ZIP)

    # ZIP 열기 + 안전 검증
    try:
        zf = zipfile.ZipFile(_io.BytesIO(data), "r")
    except zipfile.BadZipFile:
        raise HTTPException(400, "유효하지 않은 ZIP 파일")
    entries = _safe_zip_extract_paths(zf)

    # JSONL entry 찾기
    jsonl_entry = _find_jsonl_entry(entries)
    if jsonl_entry is None:
        raise HTTPException(400, "ZIP 안에 .jsonl 파일이 없습니다 (problems.jsonl 권장)")

    # 이미지 entry 수집 (확장자로 판별)
    image_entries: list[zipfile.ZipInfo] = []
    for e in entries:
        ext = Path(e.filename).suffix.lower()
        if ext in ALLOWED_IMAGE_EXTS and e is not jsonl_entry:
            image_entries.append(e)
    if len(image_entries) > MAX_IMAGES_PER_ZIP:
        raise HTTPException(
            400, f"이미지가 너무 많습니다: {len(image_entries)} > {MAX_IMAGES_PER_ZIP}",
        )

    # JSONL 파싱 + 1차 검증 (dry_run이면 여기까지만)
    try:
        raw = zf.read(jsonl_entry)
    except KeyError:
        raise HTTPException(400, "JSONL 추출 실패")
    parsed = _parse_jsonl(raw)
    if len(parsed) > MAX_JSONL_PROBLEMS:
        raise HTTPException(
            400, f"한 번에 최대 {MAX_JSONL_PROBLEMS}문제까지 import 가능",
        )

    errors: list[dict] = []
    inlines: list[tuple[int, ProblemInline]] = []
    for line_no, item in parsed:
        if isinstance(item, str):
            errors.append({"line": line_no, "message": item})
            continue
        pi, err = _validate_inline(item)
        if err or pi is None:
            errors.append({"line": line_no, "message": err or "unknown error"})
            continue
        inlines.append((line_no, pi))

    # content 안에서 참조된 이미지 키 수집
    img_ref_pattern = re.compile(r'!\[[^\]]*\]\(\s*(\.?/?[\w\-./]+\.(?:png|jpg|jpeg|webp|gif|svg))\s*\)', re.IGNORECASE)
    referenced_keys: set[str] = set()
    for _ln, pi in inlines:
        for m in img_ref_pattern.finditer(pi.content or ""):
            ref = m.group(1).lstrip("./").lstrip("/")
            referenced_keys.add(ref)

    # ZIP에 실제 있는 이미지 키 (filename 그대로)
    available_keys = {e.filename for e in image_entries}
    missing_images = sorted(referenced_keys - available_keys)
    for k in missing_images:
        errors.append({"line": 0, "message": f"이미지 누락: {k}"})

    result: dict = {
        "total": len(parsed),
        "valid": len(inlines),
        "errors": errors,
        "imported_images": 0,
        "created_problem_ids": [],
        "problem_set_id": None,
    }

    if dry_run:
        return result
    if not inlines:
        raise HTTPException(400, "유효한 문제가 한 줄도 없습니다.")
    if missing_images:
        raise HTTPException(
            400,
            f"이미지가 ZIP에 없습니다: {missing_images[:5]} (총 {len(missing_images)}건). "
            f"dry_run으로 확인 후 보완해주세요.",
        )

    # 이미지 → storage 저장
    await ensure_dir_async(COURSEWARE_STORAGE_DIR)
    image_map: dict[str, str] = {}  # 원본 키 → 새 URL
    for e in image_entries:
        if e.filename not in referenced_keys:
            # 참조 안 된 이미지는 skip (불필요한 disk 낭비 방지)
            continue
        ext = Path(e.filename).suffix.lower()
        token = secrets.token_urlsafe(12)  # 16자 url-safe
        stored_name = f"{token}{ext}"
        try:
            blob = zf.read(e)
        except KeyError:
            errors.append({"line": 0, "message": f"이미지 추출 실패: {e.filename}"})
            continue
        await write_bytes_async(COURSEWARE_STORAGE_DIR / stored_name, blob)
        image_map[e.filename] = f"/api/files/storage/courseware/{stored_name}"

    # Problem row 일괄 생성 (content URL 치환)
    new_objs: list[Problem] = []
    for _line_no, pi in inlines:
        new_content = _rewrite_image_paths(pi.content, image_map)
        obj = Problem(
            department="math",
            subject=pi.subject or "",
            difficulty=pi.difficulty,
            question_type=pi.type,
            content=new_content,
            solution=pi.solution,
            answer=pi.answer,
            answer_data=pi.answer_data,
            tags=pi.tags,
            is_visible=True,
            review_status="pending",
            created_by_id=user.id,
        )
        db.add(obj)
        new_objs.append(obj)
    await db.flush()
    created_ids = [o.id for o in new_objs]
    result["created_problem_ids"] = created_ids
    result["imported_images"] = len(image_map)

    if create_set:
        title = (set_title or "").strip() or f"ZIP Import {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ps = CourseProblemSet(
            course_id=cid,
            title=title,
            description=f"ZIP import — {len(created_ids)}문제 · 이미지 {len(image_map)}장",
            problem_ids=created_ids,
            status="draft",
            created_by=user.id,
        )
        db.add(ps)
        await db.flush()
        result["problem_set_id"] = ps.id

    await log_action(
        db, user, "courseware.problems.import_zip",
        target=f"course:{cid} count:{len(created_ids)} images:{len(image_map)}",
        detail=f"errors:{len(errors)} set:{result['problem_set_id']}",
        request=request,
    )
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 결과 export — CSV / XLSX
# ─────────────────────────────────────────────────────────────────────────────

async def _load_results_for_export(db: AsyncSession, psid: int) -> dict:
    """학생별 best_score + 문제별 정답률 + 시도 raw rows.

    구조:
      {
        "ps": CourseProblemSet,
        "problems": list[Problem],
        "attempts": list[StudentProblemAttempt],
        "students": dict[user_id, {"name": str, "snum": str|None}],
      }
    """
    ps = await db.get(CourseProblemSet, psid)
    if not ps or ps.deleted_at is not None:
        raise HTTPException(404)

    pids = ps.problem_ids or []
    problems: list[Problem] = []
    if pids:
        rows = (await db.execute(
            select(Problem).where(Problem.id.in_(pids))
        )).scalars().all()
        by_id = {p.id: p for p in rows}
        problems = [by_id[i] for i in pids if i in by_id]

    attempts = (await db.execute(
        select(StudentProblemAttempt).where(
            StudentProblemAttempt.problem_set_id == psid,
        )
    )).scalars().all()

    student_ids = list({a.student_id for a in attempts})
    students: dict[int, dict] = {}
    if student_ids:
        rows = (await db.execute(
            select(User.id, User.name, User.grade, User.class_number, User.student_number).where(
                User.id.in_(student_ids)
            )
        )).all()
        for r in rows:
            uid, name, grade, cls, num = r
            snum = None
            if grade and cls and num:
                snum = f"{grade}{int(cls):02d}{int(num):02d}"
            students[uid] = {"name": name, "snum": snum}

    return {"ps": ps, "problems": problems, "attempts": attempts, "students": students}


def _aggregate_results(data: dict) -> tuple[list[dict], list[dict]]:
    """raw → 학생별 + 문제별 집계.

    학생별: [{student_id, name, snum, best_score, attempts_count, latest_at}]
    문제별: [{problem_id, total, correct, accuracy}]
    """
    ps: CourseProblemSet = data["ps"]
    problems: list[Problem] = data["problems"]
    attempts: list[StudentProblemAttempt] = data["attempts"]
    students = data["students"]

    # 학생별 — attempt_n별 score sum의 max
    per_student: dict[int, dict] = {}
    for a in attempts:
        s = per_student.setdefault(a.student_id, {
            "student_id": a.student_id,
            "name": students.get(a.student_id, {}).get("name", f"#{a.student_id}"),
            "snum": students.get(a.student_id, {}).get("snum", ""),
            "by_attempt": {},  # attempt_n -> score sum
            "latest_at": None,
        })
        att_n = a.attempt_number
        cur = s["by_attempt"].get(att_n, 0.0)
        if a.is_correct is True:
            cur += (a.auto_score or 1.0)
        elif a.is_correct is False:
            cur += (a.auto_score or 0.0)
        elif a.manual_score is not None:
            cur += a.manual_score
        s["by_attempt"][att_n] = cur
        if a.submitted_at and (s["latest_at"] is None or a.submitted_at > s["latest_at"]):
            s["latest_at"] = a.submitted_at

    students_out = []
    for s in per_student.values():
        best = max(s["by_attempt"].values(), default=0.0)
        students_out.append({
            "student_id": s["student_id"],
            "name": s["name"],
            "snum": s["snum"],
            "best_score": round(best, 3),
            "attempts_count": len(s["by_attempt"]),
            "latest_at": s["latest_at"].isoformat() if s["latest_at"] else "",
        })
    students_out.sort(key=lambda x: (x["snum"] or "", x["name"]))

    # 문제별 정답률
    per_problem: dict[int, dict] = {}
    for a in attempts:
        p = per_problem.setdefault(a.problem_id, {
            "problem_id": a.problem_id,
            "total": 0, "correct": 0,
        })
        p["total"] += 1
        if a.is_correct is True:
            p["correct"] += 1

    problems_out = []
    for p in problems:
        d = per_problem.get(p.id, {"problem_id": p.id, "total": 0, "correct": 0})
        total = d["total"]
        problems_out.append({
            "problem_id": p.id,
            "type": p.question_type,
            "answer": p.answer or "",
            "total_submissions": total,
            "correct_count": d["correct"],
            "accuracy": round(d["correct"] / total, 3) if total > 0 else 0.0,
        })
    return students_out, problems_out


@router.get("/problem-sets/{psid}/results.csv")
async def results_csv(
    psid: int,
    user: User = Depends(require_permission("classroom.courseware.grade")),
    db: AsyncSession = Depends(get_db),
):
    """학생별 결과 CSV (UTF-8 BOM, Excel 호환)."""
    data = await _load_results_for_export(db, psid)
    course = await db.get(Course, data["ps"].course_id)
    if not course:
        raise HTTPException(404)
    if not await is_course_editor_or_admin(db, course, user):
        raise HTTPException(403)
    students, _problems = _aggregate_results(data)

    buf = _io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["학번", "이름", "시도수", "최고점수", "최근시도"])
    for s in students:
        writer.writerow([s["snum"], s["name"], s["attempts_count"], s["best_score"], s["latest_at"]])

    content = "﻿" + buf.getvalue()
    fname = f"problem-set-{psid}-results.csv"
    return StreamingResponse(
        _io.BytesIO(content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _build_xlsx(students: list[dict], problems: list[dict], ps_title: str) -> bytes:
    """openpyxl로 XLSX 작성 (sync; asyncio.to_thread로 호출)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    # 시트 1: 학생별
    ws1 = wb.active
    ws1.title = "학생별"
    headers1 = ["학번", "이름", "시도수", "최고점수", "최근시도"]
    ws1.append(headers1)
    for h in ws1[1]:
        h.font = Font(bold=True)
        h.fill = PatternFill("solid", fgColor="EEEEEE")
    for s in students:
        ws1.append([s["snum"], s["name"], s["attempts_count"], s["best_score"], s["latest_at"]])
    for col_idx in range(1, len(headers1) + 1):
        ws1.column_dimensions[chr(64 + col_idx)].width = 16

    # 시트 2: 문제별 정답률
    ws2 = wb.create_sheet("문제별")
    headers2 = ["문제ID", "유형", "정답", "응시수", "정답수", "정답률"]
    ws2.append(headers2)
    for h in ws2[1]:
        h.font = Font(bold=True)
        h.fill = PatternFill("solid", fgColor="EEEEEE")
    for p in problems:
        ws2.append([
            p["problem_id"], p["type"], p["answer"],
            p["total_submissions"], p["correct_count"], p["accuracy"],
        ])
    for col_idx in range(1, len(headers2) + 1):
        ws2.column_dimensions[chr(64 + col_idx)].width = 14

    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.get("/problem-sets/{psid}/results.xlsx")
async def results_xlsx(
    psid: int,
    user: User = Depends(require_permission("classroom.courseware.grade")),
    db: AsyncSession = Depends(get_db),
):
    """학생별 + 문제별 결과 XLSX 2시트."""
    data = await _load_results_for_export(db, psid)
    course = await db.get(Course, data["ps"].course_id)
    if not course:
        raise HTTPException(404)
    if not await is_course_editor_or_admin(db, course, user):
        raise HTTPException(403)
    students, problems = _aggregate_results(data)
    blob = await asyncio.to_thread(_build_xlsx, students, problems, data["ps"].title)

    fname = f"problem-set-{psid}-results.xlsx"
    return StreamingResponse(
        _io.BytesIO(blob),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

"""Survey 결과 + CSV export — 작성자/관리자 전용."""

import csv
import io

from fastapi import Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.permissions import require_permission
from app.models.classroom_surveys import (
    Survey, SurveyAnswer, SurveyQuestion, SurveyResponse,
)
from app.models.user import User
from app.modules.classroom_surveys._helpers import (
    can_manage, question_to_dict, survey_to_dict,
)
from app.modules.classroom_surveys.router import router


@router.get("/{sid}/results")
async def get_results(
    sid: int,
    user: User = Depends(require_permission("classroom.survey.view_results")),
    db: AsyncSession = Depends(get_db),
):
    s = await db.get(Survey, sid)
    if not s:
        raise HTTPException(404)
    if not can_manage(user, s):
        raise HTTPException(403, "본인 설문 결과만 조회 가능")

    qs = (await db.execute(
        select(SurveyQuestion).where(SurveyQuestion.survey_id == sid)
        .order_by(SurveyQuestion.order, SurveyQuestion.id)
    )).scalars().all()

    responses = (await db.execute(
        select(SurveyResponse).where(SurveyResponse.survey_id == sid)
        .order_by(desc(SurveyResponse.submitted_at))
        .limit(2000)  # 전교생(1500) 상한 — 응답 폭증 시 OOM 방어
    )).scalars().all()
    response_ids = [r.id for r in responses]
    answers = []
    if response_ids:
        answers = (await db.execute(
            select(SurveyAnswer).where(SurveyAnswer.response_id.in_(response_ids))
        )).scalars().all()

    respondent_ids = {r.respondent_id for r in responses if r.respondent_id}
    respondents: dict[int, str] = {}
    if respondent_ids and not s.is_anonymous:
        urows = (await db.execute(
            select(User).where(User.id.in_(respondent_ids))
        )).scalars().all()
        respondents = {u.id: u.name for u in urows}

    answers_by_qid: dict[int, list[SurveyAnswer]] = {}
    for a in answers:
        answers_by_qid.setdefault(a.question_id, []).append(a)

    question_summary = []
    for q in qs:
        ans = answers_by_qid.get(q.id, [])
        summary: dict = {
            **question_to_dict(q),
            "response_count": len(ans),
        }
        if q.question_type in ("single_choice", "multi_choice"):
            counts: dict[str, int] = {opt: 0 for opt in (q.options or [])}
            for a in ans:
                for v in (a.choice_values or []):
                    counts[v] = counts.get(v, 0) + 1
            summary["choice_counts"] = counts
        elif q.question_type == "rating":
            ratings = [a.rating_value for a in ans if a.rating_value is not None]
            counts = {i: 0 for i in range(1, q.rating_max + 1)}
            for r in ratings:
                if 1 <= r <= q.rating_max:
                    counts[r] += 1
            summary["rating_counts"] = counts
            summary["rating_avg"] = round(sum(ratings) / len(ratings), 2) if ratings else None
        else:
            summary["text_values"] = [a.text_value for a in ans if a.text_value]
        question_summary.append(summary)

    # 응답별 answer 매핑 — 개별 보기 탭에서 사용
    answers_by_resp: dict[int, list[dict]] = {}
    for a in answers:
        answers_by_resp.setdefault(a.response_id, []).append({
            "question_id": a.question_id,
            "text_value": a.text_value,
            "choice_values": a.choice_values,
            "rating_value": a.rating_value,
        })

    return {
        "survey": survey_to_dict(s),
        "response_count": len(responses),
        "questions": question_summary,
        "responses": [
            {
                "id": r.id,
                "respondent_id": r.respondent_id,
                "respondent_name": (
                    respondents.get(r.respondent_id) if r.respondent_id else None
                ),
                "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
                "answers": answers_by_resp.get(r.id, []),
            }
            for r in responses
        ],
    }


@router.get("/{sid}/results.csv")
async def export_results_csv(
    sid: int,
    user: User = Depends(require_permission("classroom.survey.view_results")),
    db: AsyncSession = Depends(get_db),
):
    """결과 CSV. 한 행 = 한 응답. UTF-8 BOM 포함."""
    s = await db.get(Survey, sid)
    if not s:
        raise HTTPException(404)
    if not can_manage(user, s):
        raise HTTPException(403)

    qs = (await db.execute(
        select(SurveyQuestion).where(SurveyQuestion.survey_id == sid)
        .order_by(SurveyQuestion.order, SurveyQuestion.id)
    )).scalars().all()

    responses = (await db.execute(
        select(SurveyResponse).where(SurveyResponse.survey_id == sid)
        .order_by(SurveyResponse.submitted_at)
        .limit(2000)  # 전교생(1500) 상한 — 응답 폭증 시 OOM 방어
    )).scalars().all()
    response_ids = [r.id for r in responses]
    answers = []
    if response_ids:
        answers = (await db.execute(
            select(SurveyAnswer).where(SurveyAnswer.response_id.in_(response_ids))
        )).scalars().all()

    answers_by_resp: dict[int, dict[int, SurveyAnswer]] = {}
    for a in answers:
        answers_by_resp.setdefault(a.response_id, {})[a.question_id] = a

    respondent_ids = {r.respondent_id for r in responses if r.respondent_id}
    respondents: dict[int, str] = {}
    if respondent_ids and not s.is_anonymous:
        urows = (await db.execute(
            select(User).where(User.id.in_(respondent_ids))
        )).scalars().all()
        respondents = {u.id: u.name for u in urows}

    buf = io.StringIO()
    buf.write("﻿")  # Excel UTF-8 BOM
    w = csv.writer(buf)
    header = ["응답ID", "응답자", "제출시각"]
    for q in qs:
        header.append(q.question_text)
    w.writerow(header)

    for r in responses:
        row = [
            r.id,
            "(익명)" if s.is_anonymous else respondents.get(r.respondent_id or 0, ""),
            r.submitted_at.isoformat() if r.submitted_at else "",
        ]
        ans_map = answers_by_resp.get(r.id, {})
        for q in qs:
            a = ans_map.get(q.id)
            if not a:
                row.append("")
            elif q.question_type in ("short_text", "long_text", "date"):
                row.append(a.text_value or "")
            elif q.question_type in ("single_choice", "multi_choice"):
                row.append(" | ".join(a.choice_values or []))
            elif q.question_type == "rating":
                row.append(str(a.rating_value) if a.rating_value is not None else "")
            else:
                row.append("")
        w.writerow(row)

    buf.seek(0)
    filename = f"survey_{sid}_results.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{sid}/results.xlsx")
async def export_results_xlsx(
    sid: int,
    user: User = Depends(require_permission("classroom.survey.view_results")),
    db: AsyncSession = Depends(get_db),
):
    """결과 Excel (.xlsx) — 한컴 셀·MS Excel·구글시트 모두 호환.

    한 행 = 한 응답. 첫 sheet = 응답 raw 데이터. openpyxl 사용 — CPU bound이지만
    응답 수십~수백 행이라 그대로 동기 처리 (event loop 영향 미미).
    """
    import asyncio
    from openpyxl import Workbook

    s = await db.get(Survey, sid)
    if not s:
        raise HTTPException(404)
    if not can_manage(user, s):
        raise HTTPException(403)

    qs = (await db.execute(
        select(SurveyQuestion).where(SurveyQuestion.survey_id == sid)
        .order_by(SurveyQuestion.order, SurveyQuestion.id)
    )).scalars().all()

    responses = (await db.execute(
        select(SurveyResponse).where(SurveyResponse.survey_id == sid)
        .order_by(SurveyResponse.submitted_at)
        .limit(2000)  # 전교생(1500) 상한 — 응답 폭증 시 OOM 방어
    )).scalars().all()
    response_ids = [r.id for r in responses]
    answers = []
    if response_ids:
        answers = (await db.execute(
            select(SurveyAnswer).where(SurveyAnswer.response_id.in_(response_ids))
        )).scalars().all()

    answers_by_resp: dict[int, dict[int, SurveyAnswer]] = {}
    for a in answers:
        answers_by_resp.setdefault(a.response_id, {})[a.question_id] = a

    respondent_ids = {r.respondent_id for r in responses if r.respondent_id}
    respondents: dict[int, str] = {}
    if respondent_ids and not s.is_anonymous:
        urows = (await db.execute(
            select(User).where(User.id.in_(respondent_ids))
        )).scalars().all()
        respondents = {u.id: u.name for u in urows}

    def _build_xlsx() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "응답"
        header = ["응답ID", "응답자", "제출시각"] + [q.question_text for q in qs]
        ws.append(header)
        # 헤더 굵게
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        for r in responses:
            row = [
                r.id,
                "(익명)" if s.is_anonymous else respondents.get(r.respondent_id or 0, ""),
                r.submitted_at.replace(tzinfo=None) if r.submitted_at else "",
            ]
            ans_map = answers_by_resp.get(r.id, {})
            for q in qs:
                a = ans_map.get(q.id)
                if not a:
                    row.append("")
                elif q.question_type in ("short_text", "long_text", "date"):
                    row.append(a.text_value or "")
                elif q.question_type in ("single_choice", "multi_choice"):
                    row.append(" | ".join(a.choice_values or []))
                elif q.question_type == "rating":
                    row.append(a.rating_value if a.rating_value is not None else "")
                else:
                    row.append("")
            ws.append(row)

        # 컬럼 폭 자동 (간단 추정)
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(50, max(10, max_len + 2))

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    # CPU-bound — to_thread로 event loop 비차단
    data = await asyncio.to_thread(_build_xlsx)

    filename = f"survey_{sid}_results.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
